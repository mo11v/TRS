"""
TRS Platform — Database Layer
SQLite with WAL mode + helpers
"""
import sqlite3
import logging
from pathlib import Path
from config.settings import DB_PATH, ADMIN_INIT_PASS, BASE_DIR, UPLOADS_DIR
from services.cloud_db_backup import request_db_backup

logger = logging.getLogger(__name__)

SCHEMA = """
PRAGMA journal_mode=WAL;
PRAGMA foreign_keys=ON;

-- ═══════════════════════════════════════════
-- USERS
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS users (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    username    TEXT NOT NULL UNIQUE,
    full_name   TEXT NOT NULL DEFAULT '',
    email       TEXT,
    role        TEXT NOT NULL DEFAULT 'Operator',
    password    TEXT NOT NULL,
    is_active   INTEGER NOT NULL DEFAULT 1,
    created_at  TEXT NOT NULL DEFAULT (datetime('now'))
);

-- ═══════════════════════════════════════════
-- DEVICES (Torque Turn Systems / Tong Units)
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS devices (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    code        TEXT NOT NULL UNIQUE,
    name        TEXT NOT NULL,
    category    TEXT NOT NULL DEFAULT 'Torque Turn System',
    model       TEXT,
    serial_no   TEXT,
    status      TEXT NOT NULL DEFAULT 'Available',
    location    TEXT,
    notes       TEXT,
    created_at  TEXT NOT NULL DEFAULT (datetime('now'))
);

-- ═══════════════════════════════════════════
-- JOBS
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS jobs (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    job_number      TEXT NOT NULL UNIQUE,
    customer        TEXT,
    rig             TEXT,
    well            TEXT,
    field           TEXT,
    country         TEXT,
    assigned_device_id INTEGER,
    status          TEXT NOT NULL DEFAULT 'Planned',
    start_date      TEXT,
    end_date        TEXT,
    notes           TEXT,
    created_by      INTEGER,
    created_at      TEXT NOT NULL DEFAULT (datetime('now')),
    FOREIGN KEY (assigned_device_id) REFERENCES devices(id),
    FOREIGN KEY (created_by)         REFERENCES users(id)
);

-- ═══════════════════════════════════════════
-- MTT DATA FILES
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS job_data_files (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    job_id      INTEGER NOT NULL,
    device_id   INTEGER,
    file_name   TEXT NOT NULL,
    file_path   TEXT NOT NULL,
    file_type   TEXT NOT NULL DEFAULT 'pdf',
    row_count   INTEGER DEFAULT 0,
    parsed      INTEGER NOT NULL DEFAULT 0,
    parse_error TEXT,
    uploaded_by INTEGER,
    created_at  TEXT NOT NULL DEFAULT (datetime('now')),
    FOREIGN KEY (job_id)      REFERENCES jobs(id)    ON DELETE CASCADE,
    FOREIGN KEY (device_id)   REFERENCES devices(id),
    FOREIGN KEY (uploaded_by) REFERENCES users(id)
);

-- ═══════════════════════════════════════════
-- MTT JOINT DATA (per joint record)
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS mtt_joint_data (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    job_id          INTEGER NOT NULL,
    file_id         INTEGER,
    joint_label     TEXT NOT NULL,
    joint_num       INTEGER NOT NULL,
    run_number      INTEGER DEFAULT 1,
    is_rerun        INTEGER DEFAULT 0,
    lot             TEXT,
    datetime        TEXT,
    final_torque    REAL,
    final_turns     REAL,
    shoulder_torque REAL,
    shoulder_turns  REAL,
    shoulder_rpm    REAL,
    delta_torque    REAL,
    delta_turns     REAL,
    comment         TEXT,
    created_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (job_id)  REFERENCES jobs(id)           ON DELETE CASCADE,
    FOREIGN KEY (file_id) REFERENCES job_data_files(id) ON DELETE SET NULL
);

-- ═══════════════════════════════════════════
-- MTT JOB SUMMARY (computed stats per job)
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS mtt_job_summary (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    job_id          INTEGER NOT NULL UNIQUE,
    file_id         INTEGER,
    pipe_type       TEXT,
    tong_model      TEXT,
    total_joints    INTEGER DEFAULT 0,
    ok_count        INTEGER DEFAULT 0,
    rerun_count     INTEGER DEFAULT 0,
    rerun_rate      REAL DEFAULT 0,
    outlier_count   INTEGER DEFAULT 0,
    ft_mean         REAL, ft_std REAL, ft_min REAL, ft_max REAL,
    turns_mean      REAL, rpm_mean REAL, dt_mean REAL,
    low_rpm_count   INTEGER DEFAULT 0,
    outliers_high   TEXT, outliers_low TEXT, reruns TEXT,
    fast_joints     TEXT, slow_joints TEXT,
    stats_json      TEXT,
    pdf_blob        BLOB,
    computed_at     TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE CASCADE
);

-- ═══════════════════════════════════════════
-- SCADA / LIVE READINGS (time-series)
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS scada_readings (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    device_id   INTEGER,
    job_id      INTEGER,
    ts          TEXT NOT NULL,
    torque      REAL,
    turns       REAL,
    rpm         REAL,
    pressure    REAL,
    temperature REAL,
    weight      REAL,
    source      TEXT NOT NULL DEFAULT 'live',
    created_at  TEXT NOT NULL DEFAULT (datetime('now')),
    FOREIGN KEY (device_id) REFERENCES devices(id),
    FOREIGN KEY (job_id)    REFERENCES jobs(id)
);

-- ═══════════════════════════════════════════
-- AI INSIGHTS per job
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS job_ai_insights (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    job_id            INTEGER NOT NULL UNIQUE,
    performance_score REAL DEFAULT 0,
    stability_score   REAL DEFAULT 0,
    anomaly_count     INTEGER DEFAULT 0,
    torque_trend      TEXT DEFAULT 'stable',
    insights_json     TEXT,
    analyzed_at       TEXT NOT NULL DEFAULT (datetime('now')),
    FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE CASCADE
);



-- ═══════════════════════════════════════════
-- TRS JOB FILE MANAGEMENT
-- Company / Year / Rig / Connection file tree
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS trs_job_files (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    company         TEXT NOT NULL,
    job_year        TEXT NOT NULL,
    job_month       TEXT NOT NULL DEFAULT '',
    rig             TEXT NOT NULL,
    connection_name TEXT NOT NULL,
    well_name       TEXT NOT NULL DEFAULT 'GENERAL',
    job_name        TEXT DEFAULT '',
    report_date     TEXT DEFAULT '',
    original_name   TEXT NOT NULL,
    stored_name     TEXT NOT NULL,
    file_path       TEXT NOT NULL,
    file_type       TEXT,
    file_size       INTEGER DEFAULT 0,
    file_blob       BLOB,
    notes           TEXT,
    detection_confidence INTEGER DEFAULT 0,
    detected_path   TEXT DEFAULT '',
    confirmed_by    INTEGER,
    confirmed_at    TEXT,
    uploaded_by     INTEGER,
    created_at      TEXT NOT NULL DEFAULT (datetime('now')),
    FOREIGN KEY (uploaded_by) REFERENCES users(id)
);
CREATE INDEX IF NOT EXISTS idx_trs_job_files_tree ON trs_job_files(company, job_year, job_month, rig, connection_name, well_name);

CREATE TABLE IF NOT EXISTS trs_job_folders (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    company         TEXT NOT NULL,
    job_year        TEXT NOT NULL,
    job_month       TEXT NOT NULL DEFAULT '',
    rig             TEXT NOT NULL,
    connection_name TEXT NOT NULL,
    well_name       TEXT NOT NULL DEFAULT 'GENERAL',
    notes           TEXT DEFAULT '',
    created_by      INTEGER,
    created_at      TEXT NOT NULL DEFAULT (datetime('now')),
    UNIQUE(company, job_year, job_month, rig, connection_name, well_name),
    FOREIGN KEY (created_by) REFERENCES users(id)
);
CREATE INDEX IF NOT EXISTS idx_trs_job_folders_tree ON trs_job_folders(company, job_year, rig, connection_name, well_name);

-- ═══════════════════════════════════════════
-- TRS EQUIPMENT MAINTENANCE
-- Assets + Excel-style maintenance log
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS trs_equipment (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    name              TEXT NOT NULL,
    category          TEXT DEFAULT '',
    serial_no         TEXT DEFAULT '',
    manufacture_date  TEXT,
    last_service_date TEXT,
    next_service_date TEXT,
    status            TEXT NOT NULL DEFAULT 'Available',
    location          TEXT DEFAULT '',
    notes             TEXT DEFAULT '',
    source_file_path  TEXT DEFAULT '',
    source_original_name TEXT DEFAULT '',
    created_by        INTEGER,
    created_at        TEXT NOT NULL DEFAULT (datetime('now')),
    updated_at        TEXT NOT NULL DEFAULT (datetime('now')),
    FOREIGN KEY (created_by) REFERENCES users(id)
);

CREATE TABLE IF NOT EXISTS trs_equipment_logs (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    equipment_id   INTEGER NOT NULL,
    service_date   TEXT NOT NULL,
    service_type   TEXT NOT NULL DEFAULT 'Maintenance',
    performed_by   TEXT DEFAULT '',
    next_due_date  TEXT,
    notes          TEXT DEFAULT '',
    created_by     INTEGER,
    created_at     TEXT NOT NULL DEFAULT (datetime('now')),
    FOREIGN KEY (equipment_id) REFERENCES trs_equipment(id) ON DELETE CASCADE,
    FOREIGN KEY (created_by) REFERENCES users(id)
);



-- ═══════════════════════════════════════════
-- TRS EQUIPMENT SPREADSHEET EDITOR
-- Web-editable Excel-like cells + revision log
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS trs_equipment_sheets (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    equipment_id    INTEGER NOT NULL UNIQUE,
    sheet_name      TEXT DEFAULT 'Sheet1',
    max_row         INTEGER DEFAULT 0,
    max_col         INTEGER DEFAULT 0,
    updated_by      INTEGER,
    updated_at      TEXT NOT NULL DEFAULT (datetime('now')),
    FOREIGN KEY (equipment_id) REFERENCES trs_equipment(id) ON DELETE CASCADE,
    FOREIGN KEY (updated_by) REFERENCES users(id)
);

CREATE TABLE IF NOT EXISTS trs_equipment_sheet_cells (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    equipment_id    INTEGER NOT NULL,
    row_index       INTEGER NOT NULL,
    col_index       INTEGER NOT NULL,
    value           TEXT DEFAULT '',
    updated_by      INTEGER,
    updated_at      TEXT NOT NULL DEFAULT (datetime('now')),
    UNIQUE(equipment_id, row_index, col_index),
    FOREIGN KEY (equipment_id) REFERENCES trs_equipment(id) ON DELETE CASCADE,
    FOREIGN KEY (updated_by) REFERENCES users(id)
);
CREATE INDEX IF NOT EXISTS idx_trs_equipment_cells ON trs_equipment_sheet_cells(equipment_id, row_index, col_index);



-- ═══════════════════════════════════════════
-- TRS NOTIFICATIONS + AI AUTOMATION RUNS
-- Used for multi-user live alerts and automatic recommendations
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS trs_notifications (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    level       TEXT NOT NULL DEFAULT 'info',
    title       TEXT NOT NULL,
    message     TEXT DEFAULT '',
    entity_type TEXT DEFAULT '',
    entity_id   INTEGER,
    is_read     INTEGER NOT NULL DEFAULT 0,
    created_by  INTEGER,
    created_at  TEXT NOT NULL DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_trs_notifications_read ON trs_notifications(is_read, created_at);

CREATE TABLE IF NOT EXISTS trs_ai_runs (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    run_type    TEXT NOT NULL DEFAULT 'manual',
    status      TEXT NOT NULL DEFAULT 'completed',
    summary     TEXT DEFAULT '',
    details_json TEXT DEFAULT '{}',
    created_by  INTEGER,
    created_at  TEXT NOT NULL DEFAULT (datetime('now'))
);

-- ═══════════════════════════════════════════
-- ACTIVITY LOG
-- ═══════════════════════════════════════════
CREATE TABLE IF NOT EXISTS activity_log (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    action      TEXT NOT NULL,
    user_id     INTEGER,
    entity_type TEXT,
    entity_id   INTEGER,
    detail      TEXT,
    created_at  TEXT NOT NULL DEFAULT (datetime('now'))
);

-- ═══════════════════════════════════════════
-- PERFORMANCE INDEXES
-- ═══════════════════════════════════════════
CREATE INDEX IF NOT EXISTS idx_jobs_device_status ON jobs(assigned_device_id, status, created_at);
CREATE INDEX IF NOT EXISTS idx_job_data_files_job ON job_data_files(job_id, created_at);
CREATE INDEX IF NOT EXISTS idx_mtt_joint_job ON mtt_joint_data(job_id, joint_num, run_number);
CREATE INDEX IF NOT EXISTS idx_scada_device_ts ON scada_readings(device_id, ts);
CREATE INDEX IF NOT EXISTS idx_activity_entity ON activity_log(entity_type, entity_id, created_at);
CREATE INDEX IF NOT EXISTS idx_equipment_status_next ON trs_equipment(status, next_service_date);
CREATE INDEX IF NOT EXISTS idx_equipment_logs_equipment ON trs_equipment_logs(equipment_id, service_date);
CREATE INDEX IF NOT EXISTS idx_notifications_entity ON trs_notifications(entity_type, entity_id, created_at);
"""


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH), check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def q1(sql: str, params: tuple = ()) -> dict | None:
    """Return first row as dict or None."""
    conn = get_conn()
    try:
        row = conn.execute(sql, params).fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


def qa(sql: str, params: tuple = ()) -> list[dict]:
    """Return all rows as list of dicts."""
    conn = get_conn()
    try:
        rows = conn.execute(sql, params).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def qx(sql: str, params: tuple = ()) -> int:
    """Execute INSERT/UPDATE/DELETE, return lastrowid."""
    conn = get_conn()
    try:
        cur = conn.execute(sql, params)
        conn.commit()
        last_id = cur.lastrowid
        try:
            request_db_backup(DB_PATH)
        except Exception as e:
            logger.debug(f"Cloud DB backup request skipped: {e}")
        return last_id
    finally:
        conn.close()




def seed_sample_equipment_library(conn):
    """Seed all bundled equipment Excel files into the Equipment module.

    Folder format supported:
      sample_data/equipment/HPU/111470.xlsx
      sample_data/equipment/Power Tong/110093.xlsx

    Each Excel file becomes one equipment record. Existing records are preserved;
    missing source paths/cells are refreshed only when needed.
    """
    try:
        import shutil
        from pathlib import Path
        from openpyxl import load_workbook

        sample_root = Path(BASE_DIR) / "sample_data" / "equipment"
        if not sample_root.exists():
            return

        total_seeded = 0
        total_existing = 0

        for category_dir in sorted([p for p in sample_root.iterdir() if p.is_dir()]):
            category = category_dir.name.strip() or "Equipment"
            upload_dir = Path(UPLOADS_DIR) / "equipment" / category
            upload_dir.mkdir(parents=True, exist_ok=True)

            for sample_file in sorted(category_dir.glob("*.xlsx")):
                serial_no = sample_file.stem.strip()
                if not serial_no or sample_file.name.startswith("~$"):
                    continue

                target = upload_dir / sample_file.name
                if not target.exists():
                    shutil.copy2(sample_file, target)

                row = conn.execute(
                    "SELECT id, source_file_path FROM trs_equipment WHERE category=? AND serial_no=?",
                    (category, serial_no)
                ).fetchone()

                if row:
                    equipment_id = row[0]
                    total_existing += 1
                    if not row[1]:
                        conn.execute(
                            "UPDATE trs_equipment SET source_file_path=?, source_original_name=?, updated_at=datetime('now') WHERE id=?",
                            (str(target), sample_file.name, equipment_id)
                        )
                else:
                    cur = conn.execute("""
                        INSERT INTO trs_equipment
                        (name, category, serial_no, status, location, notes, source_file_path, source_original_name, created_by)
                        VALUES (?,?,?,?,?,?,?,?,?)
                    """, (
                        f"{category} {serial_no}", category, serial_no, "Available", "Workshop",
                        f"Bundled equipment history card imported from {category}/{sample_file.name}.",
                        str(target), sample_file.name, 1
                    ))
                    equipment_id = cur.lastrowid
                    total_seeded += 1

                cell_count = conn.execute(
                    "SELECT COUNT(*) FROM trs_equipment_sheet_cells WHERE equipment_id=?", (equipment_id,)
                ).fetchone()[0]
                if cell_count:
                    continue

                try:
                    wb = load_workbook(target, data_only=False)
                    ws = wb.active
                    max_rows = min(ws.max_row or 0, 200)
                    max_cols = min(ws.max_column or 0, 30)
                    conn.execute("""
                        INSERT OR REPLACE INTO trs_equipment_sheets
                        (equipment_id, sheet_name, max_row, max_col, updated_by, updated_at)
                        VALUES (?,?,?,?,?,datetime('now'))
                    """, (equipment_id, ws.title or "Sheet1", max_rows, max_cols, 1))
                    for r in range(1, max_rows + 1):
                        for c in range(1, max_cols + 1):
                            v = ws.cell(r, c).value
                            if v is not None and str(v) != "":
                                conn.execute("""
                                    INSERT OR REPLACE INTO trs_equipment_sheet_cells
                                    (equipment_id, row_index, col_index, value, updated_by, updated_at)
                                    VALUES (?,?,?,?,?,datetime('now'))
                                """, (equipment_id, r, c, str(v), 1))
                except Exception as excel_err:
                    logger.warning(f"Equipment Excel import skipped for {sample_file}: {excel_err}")

        conn.commit()
        if total_seeded or total_existing:
            logger.info(f"Seeded equipment library: {total_seeded} new, {total_existing} existing")
    except Exception as e:
        logger.warning(f"Equipment library seed skipped: {e}")

def init_db():
    """Initialize database schema and seed admin user."""
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = get_conn()
    conn.executescript(SCHEMA)

    # Lightweight migrations for existing local databases
    def _ensure_col(table, col, ddl):
        cols = [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
        if col not in cols:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {ddl}")

    _ensure_col('trs_job_files', 'detection_confidence', 'detection_confidence INTEGER DEFAULT 0')
    _ensure_col('trs_job_files', 'detected_path', 'detected_path TEXT DEFAULT \'\'')
    _ensure_col('trs_job_files', 'confirmed_by', 'confirmed_by INTEGER')
    _ensure_col('trs_job_files', 'confirmed_at', 'confirmed_at TEXT')
    _ensure_col('trs_job_files', 'job_month', "job_month TEXT NOT NULL DEFAULT ''")
    _ensure_col('trs_job_files', 'well_name', "well_name TEXT NOT NULL DEFAULT 'GENERAL'")
    _ensure_col('trs_job_files', 'job_name', "job_name TEXT DEFAULT ''")
    _ensure_col('trs_job_files', 'report_date', "report_date TEXT DEFAULT ''")
    _ensure_col('trs_job_files', 'file_blob', 'file_blob BLOB')
    _ensure_col('trs_job_folders', 'job_month', "job_month TEXT NOT NULL DEFAULT ''")
    _ensure_col('trs_job_folders', 'well_name', "well_name TEXT NOT NULL DEFAULT 'GENERAL'")
    _ensure_col('trs_equipment', 'source_file_path', 'source_file_path TEXT DEFAULT \'\'')
    _ensure_col('trs_equipment', 'source_original_name', 'source_original_name TEXT DEFAULT \'\'')

    conn.commit()

    # Seed admin user if no users exist
    count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if count == 0:
        from services.auth_service import AuthService
        default_users = [
            ("admin", "System Admin", "Admin", ADMIN_INIT_PASS),
            ("manager", "Operations Manager", "Manager", "manager123"),
            ("engineer", "TRS Engineer", "Engineer", "engineer123"),
            ("supervisor", "Field Supervisor", "Supervisor", "supervisor123"),
            ("operator", "Field Operator Demo", "Operator", "operator123"),
        ]
        for username, full_name, role, raw_password in default_users:
            conn.execute(
                "INSERT INTO users (username, full_name, role, password) VALUES (?,?,?,?)",
                (username, full_name, role, AuthService.hash_password(raw_password))
            )
        # Seed a demo device
        conn.execute(
            "INSERT INTO devices (code, name, category, model, status) VALUES (?,?,?,?,?)",
            ("TRS-001", "TRS Unit #1", "Torque Turn System", "MTT Pro", "Available")
        )
        conn.commit()
        logger.info("TRS DB initialized — default 5-role users created")

    seed_sample_equipment_library(conn)

    conn.close()


def log_activity(action: str, user_id: int, entity_type: str = None,
                 entity_id: int = None, detail: str = None):
    qx(
        "INSERT INTO activity_log (action, user_id, entity_type, entity_id, detail) VALUES (?,?,?,?,?)",
        (action, user_id, entity_type, entity_id, detail)
    )
