"""
TRS Platform
TRS Operations | Job Files | Equipment | Torque Analysis
Run: uvicorn app:app --reload --host 0.0.0.0 --port 8000
"""

import os, logging, json, uuid, io, math, csv, re, zipfile, html, shutil
from datetime import datetime
from fastapi import FastAPI, Request, Form, UploadFile, File, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, Response, JSONResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
import uvicorn

from config.settings import (
    APP_NAME, APP_VERSION, UPLOADS_DIR, SESSION_HOURS,
    DEVICE_CATEGORIES, USER_ROLES, JOB_STATUSES, DEVICE_STATUSES, MAX_UPLOAD_MB
)
from database.db import (
    DB_PATH, get_conn, init_db, q1, qa, qx, log_activity
)
from services.auth_service import AuthService, LoginRateLimiter
from services.rbac import rbac, PERMISSIONS
from services.firestore_service import (
    fs_enabled, sync_job_to_fs, sync_summary_to_fs,
    sync_scada_to_fs, sync_device_to_fs,
    restore_from_firestore_to_sqlite,
)
from services.cloudinary_service import (
    cl_enabled, upload_mtt_original, upload_report_pdf
)
from services.cloud_db_backup import restore_db_from_cloud, backup_now
from services.google_drive_backup import start_daily_scheduler as start_google_drive_backup_scheduler

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── Bootstrap ─────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
os.makedirs(STATIC_DIR, exist_ok=True)
os.makedirs(str(UPLOADS_DIR), exist_ok=True)

app = FastAPI(title=APP_NAME, version=APP_VERSION)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

login_limiter = LoginRateLimiter()
session_store: dict[str, dict] = {}

# ══════════════════════════════════════════════════════════════
#  WEBSOCKET MANAGER — Real-time SCADA broadcast  
# ══════════════════════════════════════════════════════════════

class WSManager:
    def __init__(self):
        self.rooms: dict = {"all": []}
        self.presence: dict = {}

    async def connect(self, ws, channel="all", user=None):
        await ws.accept()
        self.rooms.setdefault(channel, []).append(ws)
        self.presence[id(ws)] = {
            "channel": channel,
            "user": (user or {}).get("username", "guest"),
            "full_name": (user or {}).get("full_name", ""),
            "connected_at": datetime.now().isoformat(timespec="seconds"),
        }
        # Only broadcast presence to the 'all' channel — not to data channels
        # This prevents job_files/equipment pages from reloading on every connection
        if channel not in ('job_files','equipment','scada'):
            await self.broadcast({"type":"presence", "event":"join", "channel":channel, "clients":self.client_count()}, channel)

    def disconnect(self, ws, channel="all"):
        room = self.rooms.get(channel, [])
        if ws in room:
            room.remove(ws)
        self.presence.pop(id(ws), None)

    async def broadcast(self, data: dict, channel="all"):
        import json
        msg = json.dumps(data, default=str)
        targets = list(set(self.rooms.get(channel, []) + self.rooms.get("all", [])))
        dead = []
        for ws in targets:
            try:
                await ws.send_text(msg)
            except Exception:
                dead.append(ws)
        for ws in dead:
            for room in self.rooms.values():
                if ws in room:
                    room.remove(ws)

    def client_count(self):
        return sum(len(v) for v in self.rooms.values())

    def status(self):
        return {
            "clients": self.client_count(),
            "rooms": {k: len(v) for k, v in self.rooms.items()},
            "presence": list(self.presence.values()),
        }

ws_manager = WSManager()




# ══════════════════════════════════════════════════════════════
#  Auth Helpers
# ══════════════════════════════════════════════════════════════

def get_user(request: Request) -> dict | None:
    token = request.cookies.get("trs_session") or request.cookies.get("jam_session")
    if not token:
        return None
    sess = session_store.get(token)
    if not sess:
        return None
    from datetime import timedelta
    exp = datetime.fromisoformat(sess.get("expires", "1970-01-01"))
    if datetime.now() > exp:
        session_store.pop(token, None)
        return None
    # Sliding window
    sess["expires"] = (datetime.now() + timedelta(hours=SESSION_HOURS)).isoformat()
    return sess.get("user")


def require_user(request: Request):
    user = get_user(request)
    if not user:
        return None, RedirectResponse("/login", status_code=303)
    return user, None


def require_role(request: Request, roles: list[str]):
    user, redir = require_user(request)
    if redir:
        return None, redir
    if user["role"] not in roles:
        return None, RedirectResponse("/", status_code=303)
    return user, None


def render(request: Request, name: str, ctx: dict):
    ctx.setdefault("app_name", APP_NAME)
    ctx.setdefault("app_version", APP_VERSION)
    u = get_user(request)
    ctx.setdefault("user", u)
    ctx.setdefault("now", datetime.now().strftime("%Y-%m-%d %H:%M"))
    # Inject user permissions into every template
    if u:
        role = u.get("role", "Operator")
        role_for_perms = "Operator" if role == "Field Operator" else role
        ctx["perms"] = [p for p, roles in PERMISSIONS.items() if role_for_perms in roles]
        ctx["user_role"] = role_for_perms
    else:
        ctx["perms"] = []
        ctx["user_role"] = ""
    # Starlette/FastAPI newer versions expect: TemplateResponse(request, name, context)
    # Old order caused: TypeError: unhashable type: 'dict' on /login
    return templates.TemplateResponse(request, name, ctx)


# ══════════════════════════════════════════════════════════════
#  Startup / Shutdown
# ══════════════════════════════════════════════════════════════

@app.on_event("startup")
async def startup():
    # ── Create dirs (important on Render /tmp) ────────────────
    import pathlib
    for d in [UPLOADS_DIR, str(pathlib.Path(str(DB_PATH)).parent)]:
        try:
            pathlib.Path(str(d)).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            logger.warning(f"mkdir failed: {d} — {e}")

    # On Render/free hosts the local disk may reset. Restore the latest SQLite backup
    # from Cloudinary before initializing schema so Excel edits and activity log persist.
    try:
        restore_db_from_cloud(DB_PATH)
    except Exception as e:
        logger.warning(f"Cloud DB restore skipped: {e}")

    init_db()
    try:
        ensure_hpu_sample_present()
    except Exception as e:
        logger.warning(f"HPU sample check skipped: {e}")
    logger.info(f"TRS Platform v{APP_VERSION} started")
    logger.info(f"DB:      {DB_PATH}")
    logger.info(f"Uploads: {UPLOADS_DIR}")
    logger.info(f"ESP Key: {ESP_API_KEY[:8]}...")

    # Restore from Firestore if SQLite empty (Render restart)
    try:
        restored = restore_from_firestore_to_sqlite()
        if restored > 0:
            logger.info(f"Restored {restored} jobs from Firestore")
        elif fs_enabled():
            logger.info("Firestore connected - sync active")
    except Exception as e:
        logger.warning(f"Firestore restore skipped: {e}")

    if cl_enabled():
        logger.info("Cloudinary connected - PDF storage active")
        try:
            backup_now(DB_PATH)
        except Exception as e:
            logger.warning(f"Initial cloud DB backup skipped: {e}")

    # Optional Google Drive daily ZIP backup. Enable with TRS_GOOGLE_DRIVE_BACKUP=1
    try:
        start_google_drive_backup_scheduler(base_dir=BASE_DIR, db_path=DB_PATH)
    except Exception as e:
        logger.warning(f"Google Drive backup scheduler skipped: {e}")


# ══════════════════════════════════════════════════════════════
#  AUTH
# ══════════════════════════════════════════════════════════════

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    if get_user(request):
        return RedirectResponse("/", status_code=303)
    return render(request, "login.html", {"error": None})


@app.post("/login")
async def login_post(request: Request,
                     username: str = Form(...),
                     password: str = Form(...)):
    ip = request.client.host if request.client else "local"
    blocked, wait = login_limiter.is_blocked(ip)
    if blocked:
        minutes = max(1, (wait + 59) // 60)
        return render(request, "login.html", {"error": f"Too many attempts. Try in {minutes} minutes."})

    username = username.strip()
    user = q1("SELECT * FROM users WHERE username=? AND is_active=1", (username,))
    if not user or not AuthService.verify_password(password, user["password"]):
        login_limiter.record_attempt(ip)
        return render(request, "login.html", {"error": "Invalid username or password"})

    login_limiter.clear(ip)
    token = AuthService.issue_session_token()
    from datetime import timedelta
    session_store[token] = {
        "user": dict(user),
        "expires": (datetime.now() + timedelta(hours=SESSION_HOURS)).isoformat()
    }
    log_activity("login", user["id"], "user", user["id"], f"Login from {ip}")
    resp = RedirectResponse("/", status_code=303)
    resp.set_cookie("trs_session", token, httponly=True, samesite="lax",
                    max_age=SESSION_HOURS * 3600)
    return resp


@app.get("/logout")
async def logout(request: Request):
    token = request.cookies.get("trs_session") or request.cookies.get("jam_session")
    if token:
        session_store.pop(token, None)
    resp = RedirectResponse("/login", status_code=303)
    resp.delete_cookie("trs_session")
    resp.delete_cookie("jam_session")
    return resp


# ══════════════════════════════════════════════════════════════
#  DASHBOARD
# ══════════════════════════════════════════════════════════════

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    user, redir = require_user(request)
    if redir: return redir

    role = "Operator" if user.get("role") == "Field Operator" else user.get("role", "Operator")

    # Base stats (all roles)
    total_jobs    = q1("SELECT COUNT(*) c FROM jobs")["c"]
    active_jobs   = q1("SELECT COUNT(*) c FROM jobs WHERE status='Active'")["c"]
    total_devices = q1("SELECT COUNT(*) c FROM devices")["c"]
    avail_devices = q1("SELECT COUNT(*) c FROM devices WHERE status='Available'")["c"]
    total_joints  = q1("SELECT COALESCE(SUM(total_joints),0) c FROM mtt_job_summary")["c"]
    analyzed_jobs = q1("SELECT COUNT(*) c FROM mtt_job_summary")["c"]
    total_users   = q1("SELECT COUNT(*) c FROM users WHERE is_active=1")["c"] if role == "Admin" else 0
    job_file_count = q1("SELECT COUNT(*) c FROM trs_job_files")["c"]
    trs_equipment_count = q1("SELECT COUNT(*) c FROM trs_equipment")["c"]
    trs_equipment_overdue = q1("SELECT COUNT(*) c FROM trs_equipment WHERE next_service_date IS NOT NULL AND date(next_service_date) < date('now')")["c"]
    trs_equipment_due_soon = q1("SELECT COUNT(*) c FROM trs_equipment WHERE next_service_date IS NOT NULL AND date(next_service_date) BETWEEN date('now') AND date('now','+30 day')")["c"]
    recent_job_files = qa("""
        SELECT f.*, u.username, u.full_name
        FROM trs_job_files f LEFT JOIN users u ON u.id=f.uploaded_by
        ORDER BY f.created_at DESC LIMIT 6
    """)
    equipment_alerts = qa("""
        SELECT id, name, category, serial_no, next_service_date,
               CASE
                 WHEN next_service_date IS NOT NULL AND date(next_service_date) < date('now') THEN 'OVERDUE'
                 WHEN next_service_date IS NOT NULL AND date(next_service_date) <= date('now','+30 day') THEN 'DUE_SOON'
                 ELSE 'OK'
               END AS service_state
        FROM trs_equipment
        WHERE next_service_date IS NOT NULL AND date(next_service_date) <= date('now','+30 day')
        ORDER BY date(next_service_date) ASC LIMIT 6
    """)

    # Recent jobs (Field Operator sees only active)
    if role == "Operator":
        recent_jobs = qa("""
            SELECT j.id, j.job_number, j.customer, j.rig, j.status, j.start_date,
                   d.code device_code, d.name device_name,
                   m.total_joints, m.ft_mean, m.rerun_rate
            FROM jobs j
            LEFT JOIN devices d ON d.id=j.assigned_device_id
            LEFT JOIN mtt_job_summary m ON m.job_id=j.id
            WHERE j.status='Active'
            ORDER BY j.created_at DESC LIMIT 10
        """)
    else:
        recent_jobs = qa("""
            SELECT j.id, j.job_number, j.customer, j.rig, j.status, j.start_date,
                   d.code device_code, d.name device_name,
                   m.total_joints, m.ft_mean, m.rerun_rate, m.computed_at
            FROM jobs j
            LEFT JOIN devices d ON d.id=j.assigned_device_id
            LEFT JOIN mtt_job_summary m ON m.job_id=j.id
            ORDER BY j.created_at DESC LIMIT 8
        """)

    # Fleet chart (Engineer and above)
    fleet_chart = []
    if role in ("Admin", "Manager", "Engineer", "Supervisor", "Operator"):
        fleet_chart = qa("""
            SELECT j.job_number, m.ft_mean, m.ft_min, m.ft_max, m.rerun_rate
            FROM mtt_job_summary m JOIN jobs j ON j.id=m.job_id
            ORDER BY m.computed_at DESC LIMIT 10
        """)

    # Fleet health insights
    fleet_health = []
    if role in ("Admin", "Manager", "Engineer", "Supervisor", "Operator") and fleet_chart:
        all_rerun = [j["rerun_rate"] for j in fleet_chart if j["rerun_rate"] is not None]
        all_ft = [j["ft_mean"] for j in fleet_chart if j["ft_mean"] is not None]
        avg_rerun = round(sum(all_rerun)/len(all_rerun), 1) if all_rerun else 0
        if avg_rerun > 10:
            fleet_health.append({"type":"critical","icon":"🔴","title":f"High Fleet Rerun Rate: {avg_rerun}%","detail":"Multiple jobs showing elevated rerun rates — check equipment condition"})
        elif avg_rerun > 5:
            fleet_health.append({"type":"warning","icon":"🟡","title":f"Elevated Rerun Rate: {avg_rerun}%","detail":"Monitor rerun trends across the fleet"})
        else:
            fleet_health.append({"type":"success","icon":"🟢","title":f"Good Fleet Rerun Rate: {avg_rerun}%","detail":"Fleet rerun rates within acceptable range"})

        if all_ft:
            mean_ft = sum(all_ft)/len(all_ft)
            spread = max(all_ft) - min(all_ft)
            if spread > mean_ft * 0.3:
                fleet_health.append({"type":"warning","icon":"📊","title":"High Torque Variance Across Fleet","detail":f"Range: {int(min(all_ft))}–{int(max(all_ft))} ft·lb — check pipe grade consistency"})
            else:
                fleet_health.append({"type":"success","icon":"🎯","title":"Consistent Fleet Torque","detail":f"Fleet avg: {int(mean_ft)} ft·lb — stable across jobs"})

        busy_devices = q1("SELECT COUNT(*) c FROM devices WHERE status='In Job'")["c"]
        if busy_devices > 0:
            fleet_health.append({"type":"info","icon":"🔧",
                "title":f"{busy_devices} Device(s) In Job",
                "detail":f"{avail_devices} available — {total_devices - avail_devices - busy_devices} in maintenance"})

    # Activity (Admin only)
    activity = []
    if role == "Admin":
        activity = qa("""
            SELECT a.*, u.username, u.full_name
            FROM activity_log a LEFT JOIN users u ON u.id=a.user_id
            ORDER BY a.created_at DESC LIMIT 10
        """)

    # Devices list for new job modal
    devices = qa("SELECT id, code, name FROM devices WHERE status='Available' ORDER BY code")

    # Role-specific performance blocks.
    operator_panel = {}
    if role in ("Operator", "Field Operator"):
        uid = user["id"]
        operator_jobs = qa("""
            SELECT j.id, j.job_number, j.customer, j.rig, j.well, j.status, j.start_date,
                   m.total_joints, m.ft_mean, m.rerun_rate, m.computed_at
            FROM jobs j
            LEFT JOIN mtt_job_summary m ON m.job_id=j.id
            WHERE j.created_by=?
            ORDER BY j.created_at DESC LIMIT 12
        """, (uid,))
        operator_uploads = qa("""
            SELECT id, original_name, company, job_year, job_month, rig, connection_name, detection_confidence, created_at
            FROM trs_job_files
            WHERE uploaded_by=?
            ORDER BY created_at DESC LIMIT 12
        """, (uid,))
        op_total = q1("SELECT COUNT(*) c FROM jobs WHERE created_by=?", (uid,))["c"]
        op_analyzed = q1("""
            SELECT COUNT(*) c
            FROM jobs j JOIN mtt_job_summary m ON m.job_id=j.id
            WHERE j.created_by=?
        """, (uid,))["c"]
        op_uploads = q1("SELECT COUNT(*) c FROM trs_job_files WHERE uploaded_by=?", (uid,))["c"]
        avg_rerun_row = q1("""
            SELECT AVG(m.rerun_rate) avg_rerun
            FROM jobs j JOIN mtt_job_summary m ON m.job_id=j.id
            WHERE j.created_by=? AND m.rerun_rate IS NOT NULL
        """, (uid,))
        avg_rerun = float(avg_rerun_row["avg_rerun"] or 0)
        score = 100
        score -= min(avg_rerun * 3, 35)
        if op_total:
            score += min((op_analyzed / max(op_total, 1)) * 10, 10)
        score = max(0, min(100, round(score)))
        operator_panel = {
            "jobs": operator_jobs,
            "uploads": operator_uploads,
            "total_jobs": op_total,
            "analyzed_jobs": op_analyzed,
            "uploaded_files": op_uploads,
            "avg_rerun": round(avg_rerun, 1),
            "score": score,
        }

    role_summary = {
        "Admin": "Full system control: users, settings, all modules and audit logs.",
        "Manager": "Operational management: jobs, files, equipment, analytics and team overview.",
        "Engineer": "Engineering workspace: analysis, AI insights, reports and technical follow-up.",
        "Supervisor": "Field supervision: team jobs, job files, equipment status and live operations.",
        "Operator": "Personal field panel: your jobs, your uploads, analysis, score and performance.",
        "Field Operator": "Personal field panel: your jobs, your uploads, analysis, score and performance.",
    }.get(role, "TRS operational dashboard.")

    return render(request, "dashboard.html", {
        "active": "dashboard",
        "total_jobs": total_jobs, "active_jobs": active_jobs,
        "total_devices": total_devices, "avail_devices": avail_devices,
        "total_joints": total_joints, "analyzed_jobs": analyzed_jobs,
        "total_users": total_users,
        "job_file_count": job_file_count, "trs_equipment_count": trs_equipment_count,
        "trs_equipment_overdue": trs_equipment_overdue, "trs_equipment_due_soon": trs_equipment_due_soon,
        "recent_job_files": recent_job_files, "equipment_alerts": equipment_alerts,
        "recent_jobs": recent_jobs, "fleet_chart": fleet_chart,
        "fleet_health": fleet_health, "activity": activity,
        "devices": devices,
        "operator_panel": operator_panel,
        "role_summary": role_summary,
    })




def safe_name(value: str, fallback: str = "item") -> str:
    """Keep uploaded-file paths predictable and safe."""
    import re
    v = (value or "").strip()
    v = re.sub(r"[^A-Za-z0-9._\- ]+", "_", v)
    v = re.sub(r"\s+", "_", v).strip("._-")
    return v or fallback


def file_ext(name: str) -> str:
    return os.path.splitext(name or "")[1].lower().lstrip(".") or "file"


def resolve_equipment_source_path(eq: dict | None) -> str:
    """Return a valid source Excel path even if the DB still contains an old absolute path."""
    if not eq:
        return ""
    current = (eq.get("source_file_path") or "").strip()
    candidates = []
    if current:
        candidates.append(current)
        candidates.append(os.path.join(BASE_DIR, current))
        candidates.append(os.path.join(str(UPLOADS_DIR), "equipment", safe_name(eq.get("category") or "HPU"), os.path.basename(current)))
    original = (eq.get("source_original_name") or "").strip()
    serial = (eq.get("serial_no") or "").strip()
    category = safe_name(eq.get("category") or "HPU")
    for name in [original, f"{serial}.xlsx", f"{serial}.xlsm", f"{serial}.xls", f"{serial}.csv"]:
        if name:
            candidates.extend([
                os.path.join(str(UPLOADS_DIR), "equipment", category, name),
                os.path.join(BASE_DIR, "uploads", "equipment", category, name),
                os.path.join(BASE_DIR, "sample_data", "equipment", category, name),
            ])
    for c in candidates:
        try:
            if c and os.path.exists(c):
                return os.path.abspath(c)
        except Exception:
            pass
    return current


def ensure_hpu_sample_present():
    """Keep HPU/111470.xlsx usable after unzipping on any PC."""
    dst_dir = os.path.join(str(UPLOADS_DIR), "equipment", "HPU")
    os.makedirs(dst_dir, exist_ok=True)
    dst = os.path.join(dst_dir, "111470.xlsx")
    src_candidates = [
        os.path.join(BASE_DIR, "sample_data", "equipment", "HPU", "111470.xlsx"),
        os.path.join(BASE_DIR, "uploads", "equipment", "HPU", "111470.xlsx"),
    ]
    if not os.path.exists(dst):
        for src in src_candidates:
            if os.path.exists(src):
                try:
                    shutil.copy2(src, dst)
                    break
                except Exception:
                    pass
    if os.path.exists(dst):
        eq = q1("SELECT * FROM trs_equipment WHERE category='HPU' AND serial_no='111470'")
        if eq:
            if resolve_equipment_source_path(eq) != os.path.abspath(dst):
                qx("UPDATE trs_equipment SET source_file_path=?, source_original_name=?, updated_at=datetime('now') WHERE id=?", (os.path.abspath(dst), "111470.xlsx", eq["id"]))
        else:
            eid = qx("""
                INSERT INTO trs_equipment (name, category, serial_no, status, location, notes, source_file_path, source_original_name)
                VALUES (?,?,?,?,?,?,?,?)
            """, ("HPU 111470", "HPU", "111470", "Available", "", "Sample HPU Excel file for testing", os.path.abspath(dst), "111470.xlsx"))
            import_equipment_sheet_from_file(eid, dst, None)


def _xlsx_color(color_obj, default=""):
    try:
        if not color_obj:
            return default
        rgb = getattr(color_obj, "rgb", None)
        if rgb and isinstance(rgb, str):
            rgb = rgb[-6:]
            if rgb and rgb != "000000":
                return f"#{rgb}"
    except Exception:
        pass
    return default




def _is_checkbox_like(value) -> bool:
    """Detect cells that should behave like web checkboxes."""
    if isinstance(value, bool):
        return True
    txt = str(value or "").strip().lower()
    return txt in {"true", "false", "yes", "no", "y", "n", "done", "not done", "☑", "☒", "✓", "✔", "x", "✗", "☐", "checked", "unchecked"}


def _checkbox_state(value) -> bool:
    if isinstance(value, bool):
        return value
    txt = str(value or "").strip().lower()
    return txt in {"true", "yes", "y", "done", "☑", "☒", "✓", "✔", "x", "checked"}


def _checkbox_save_value(checked: bool, original_value=None) -> str:
    """Keep the same style of checkbox value where possible."""
    if isinstance(original_value, bool):
        return "TRUE" if checked else "FALSE"
    txt = str(original_value or "").strip()
    low = txt.lower()
    if txt in {"☑", "☐", "☒", "✓", "✔", "✗"}:
        return "☑" if checked else "☐"
    if low in {"yes", "no", "y", "n"}:
        return "YES" if checked else "NO"
    if low in {"done", "not done"}:
        return "DONE" if checked else "NOT DONE"
    if low in {"checked", "unchecked"}:
        return "checked" if checked else "unchecked"
    return "TRUE" if checked else "FALSE"


def write_equipment_cell_to_source(equipment_id: int, row: int, col: int, value: str) -> tuple[bool, str]:
    """Write one edited web cell back into the original Excel file while preserving workbook styling/images/formulas elsewhere."""
    eq = q1("SELECT * FROM trs_equipment WHERE id=?", (equipment_id,))
    source = resolve_equipment_source_path(eq)
    if not eq or not source or not os.path.exists(source):
        return False, "Source Excel file not found"
    if file_ext(source) not in ["xlsx", "xlsm", "xltx", "xltm"]:
        return False, "Source file is not an editable XLSX/XLSM workbook"
    try:
        from openpyxl import load_workbook
        from openpyxl.cell.cell import MergedCell
        wb = load_workbook(source, keep_vba=(file_ext(source) == "xlsm"))
        ws = wb.active
        target = ws.cell(row=row, column=col)
        if isinstance(target, MergedCell):
            return False, "Cannot edit a non-primary merged cell"
        old_value = target.value
        if _is_checkbox_like(old_value) or str(value).strip().upper() in {"TRUE", "FALSE", "☑", "☐"}:
            target.value = _checkbox_state(value)
        else:
            # Preserve formulas unless the user intentionally types a formula beginning with =.
            target.value = value
        wb.save(source)
        qx("UPDATE trs_equipment SET source_file_path=?, updated_at=datetime('now') WHERE id=?", (os.path.abspath(source), equipment_id))
        return True, "Saved to original Excel"
    except Exception as e:
        write_log(f"Excel source save error equipment={equipment_id} cell={row},{col}: {e}")
        return False, str(e)

def render_styled_equipment_sheet_html(equipment_id: int, eq: dict, max_rows: int = 160, max_cols: int = 40) -> str:
    """Render an Excel-like editable HTML table preserving basic formatting from the source file."""
    source = resolve_equipment_source_path(eq)
    if not source or file_ext(source) not in ["xlsx", "xlsm", "xltx", "xltm"] or not os.path.exists(source):
        return ""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return ""
    try:
        wb = load_workbook(source, data_only=False)
        ws = wb.active
    except Exception:
        return ""
    db_cells = {(int(c["row_index"]), int(c["col_index"])): (c.get("value") or "") for c in qa("SELECT row_index,col_index,value FROM trs_equipment_sheet_cells WHERE equipment_id=?", (equipment_id,))}
    mr = min(ws.max_row or 1, max_rows); mc = min(ws.max_column or 1, max_cols)
    merged_top = {}; merged_skip=set()
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        if min_row > mr or min_col > mc: continue
        rs=min(max_row,mr)-min_row+1; cs=min(max_col,mc)-min_col+1
        merged_top[(min_row,min_col)]=(rs,cs)
        for rr in range(min_row,min(max_row,mr)+1):
            for cc in range(min_col,min(max_col,mc)+1):
                if (rr,cc)!=(min_row,min_col): merged_skip.add((rr,cc))
    logo_html=""
    try:
        imgs=getattr(ws,"_images",[]) or []
        if imgs:
            img_dir=os.path.join(STATIC_DIR,"excel_previews"); os.makedirs(img_dir,exist_ok=True)
            ext=getattr(imgs[0],"format","png") or "png"
            if ext.lower()=="jpeg": ext="jpg"
            img_name=f"equipment_{equipment_id}_logo.{ext}"; img_path=os.path.join(img_dir,img_name)
            if not os.path.exists(img_path):
                with open(img_path,"wb") as f: f.write(imgs[0]._data())
            logo_html=f'<img src="/static/excel_previews/{img_name}" class="excel-logo" alt="logo">'
    except Exception: logo_html=""
    colgroup=['<col style="width:46px">']
    for c in range(1,mc+1):
        width=ws.column_dimensions[get_column_letter(c)].width or 12
        px=max(70,min(260,int(width*7))); colgroup.append(f'<col style="width:{px}px">')
    out=['<table class="excel-preview-table" id="sheetTable">','<colgroup>',*colgroup,'</colgroup>']
    out.append('<thead><tr><th class="xl-corner">#</th>')
    for c in range(1,mc+1): out.append(f'<th class="xl-head">{excel_col_name(c)}</th>')
    out.append('</tr></thead><tbody>')
    for r in range(1,mr+1):
        height=ws.row_dimensions[r].height; tr_style=f' style="height:{int(height*1.33)}px"' if height else ''
        out.append(f'<tr{tr_style}><td class="row-num">{r}</td>')
        for c in range(1,mc+1):
            if (r,c) in merged_skip: continue
            cell=ws.cell(r,c); raw=db_cells.get((r,c),cell.value); value="" if raw is None else str(raw); styles=[]
            bg=_xlsx_color(getattr(cell.fill,"fgColor",None),"")
            if bg: styles.append(f"background:{bg}")
            fg=_xlsx_color(getattr(cell.font,"color",None),"")
            if fg: styles.append(f"color:{fg}")
            if getattr(cell.font,"bold",False): styles.append("font-weight:700")
            if getattr(cell.font,"italic",False): styles.append("font-style:italic")
            if getattr(cell.font,"sz",None): styles.append(f"font-size:{max(10,int(cell.font.sz))}px")
            align=getattr(cell.alignment,"horizontal",None)
            if align: styles.append(f"text-align:{align}")
            valign=getattr(cell.alignment,"vertical",None)
            if valign: styles.append(f"vertical-align:{valign}")
            styles.append("white-space:pre-wrap" if getattr(cell.alignment,"wrap_text",False) else "white-space:pre")
            rs,cs=merged_top.get((r,c),(1,1)); span=(f' rowspan="{rs}"' if rs>1 else '')+(f' colspan="{cs}"' if cs>1 else '')
            original_value = cell.value
            if _is_checkbox_like(raw):
                checked = _checkbox_state(raw)
                save_val = _checkbox_save_value(checked, original_value)
                content = f'<button type="button" class="xl-checkbox" data-value="{html.escape(save_val)}" aria-label="checkbox">{"☑" if checked else "☐"}</button>'
                editable = 'false'
                extra_class = ' checkbox-cell'
            else:
                content=html.escape(value)
                editable = 'true'
                extra_class = ''
            if r==1 and c==1 and logo_html: content=logo_html+content
            out.append(f'<td class="sheet-cell xl-cell{extra_class}" contenteditable="{editable}" data-row="{r}" data-col="{c}" data-original="{html.escape(str(original_value or ""))}"{span} style="{";".join(styles)}">{content}</td>')
        out.append('</tr>')
    out.append('</tbody></table>')
    return ''.join(out)


KNOWN_COMPANIES = ["Khalda", "AGIBA", "Simetar"]
KNOWN_CONNECTIONS = [
    "VAM TOP", "VAMTOP", "PH-6", "PH6", "PH 6", "PH-7", "PH7", "PH 7",
    "BTC", "LTC", "STC", "EUE", "NUE", "BUTTRESS", "TENARIS", "HYDRIL",
    "XT", "PREMIUM", "VAGT", "VAGT TOP", "HUNTING", "FOX", "TMK", "SEAL-LOCK"
]

def _norm_token(text: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", (text or "").upper())

def _clean_field(v: str) -> str:
    v = re.sub(r"\s+", " ", (v or "").strip())
    return v.strip(" -_:/\\")

def _normalize_rig(prefix: str, num: str) -> str:
    return f"{prefix.upper()} {int(num)}" if str(num).isdigit() else f"{prefix.upper()} {num}"

def _normalize_well(v: str) -> str:
    v = _clean_field(v)
    v = re.sub(r"\s+", "", v)
    v = re.sub(r"_?REPORT.*$", "", v, flags=re.I)
    return v.upper().strip("._-")

def _normalize_connection(v: str) -> str:
    raw = _clean_field(v).upper().replace("_", " ").replace("/", " ")
    raw = re.sub(r"\s+", " ", raw)
    compact = _norm_token(raw)
    if compact == "VAMTOP" or "VAMTOP" in compact:
        return "VAM TOP"
    m = re.search(r"\bPH\s*[- ]?\s*(\d{1,3})\b", raw)
    if m:
        return f"PH-{m.group(1)}"
    for c in ["BTC", "LTC", "STC", "EUE", "NUE", "BUTTRESS", "TENARIS", "HYDRIL", "XT", "PREMIUM"]:
        if c in raw or c in compact:
            return c
    return raw or "UNKNOWN CONNECTION"

def extract_pdf_text_sample(data: bytes, max_pages: int = 2) -> str:
    """Best-effort PDF text extraction for auto-classification; safe to fail."""
    if not data:
        return ""
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(data))
        parts = []
        for page in reader.pages[:max_pages]:
            try:
                parts.append(page.extract_text() or "")
            except Exception:
                pass
        return "\n".join(parts)[:12000]
    except Exception:
        return ""


def _strip_connection_specs(line: str) -> str:
    """Return the connection/thread name exactly from PDF field, minus size/grade/weight specs."""
    v = _clean_field(line).upper().replace("_", " ")
    v = re.sub(r"\s+", " ", v)
    # Fix common OCR/spelling variants but keep the PDF wording as the base.
    v = v.replace("TENARS", "TENARIS")
    v = v.replace("TENARI ", "TENARIS ")
    v = re.sub(r"\bVAM\s*TOP\b", "VAM TOP", v)
    v = re.sub(r"\bPH\s*[- ]?\s*(\d{1,3})\b", r"PH-\1", v)

    # Remove labels if the captured text still includes them.
    v = re.sub(r"^(LOT\(S\) USED|LOTS USED|LOT USED|CONNECTION USED)\s*[:\-]?\s*", "", v, flags=re.I)

    # Drop anything after another report field leaked into the line.
    v = re.split(r"\b(UNITS|COMPANY|COMPANY REP|LOCATION|LEASE|WELL|CONTRACTOR|RIG|TECH|FOOTAGE|QUANTITY|TONG TYPE|START DATE|END DATE|PARAMETER|MAXIMUM TORQUE|OPTIMUM TORQUE|MINIMUM TORQUE)\b", v)[0]
    v = _clean_field(v)

    # Remove common numeric specs: sizes, weights, torque numbers, grades.
    tokens = []
    for tok in v.split():
        t = tok.strip(" ,;:()[]{}")
        tu = t.upper()
        # Stop at first size/grade/weight/spec token; connection names are normally before these.
        if re.fullmatch(r"\d+(?:\.\d+)?", tu):
            break
        if re.fullmatch(r"\d+(?:\.\d+)?\"", tu):
            break
        if re.fullmatch(r"[A-Z]?\d{1,3}(?:\.\d+)?", tu) and re.search(r"\d", tu):
            # L80, L.80, P110, 9.2, 3500, 3.5 etc.
            break
        if re.fullmatch(r"[A-Z]\.?\d{1,3}", tu):
            break
        if tu in {"FT", "LBS", "FT-LBS", "LB", "L80", "L.80", "P110", "J55", "N80", "K55", "Q125"}:
            break
        tokens.append(t)

    cleaned = " ".join(tokens).strip(" -_:/\\")
    cleaned = re.sub(r"\s+", " ", cleaned)

    # Some reports repeat the same connection name before specs, e.g.
    # "VAM TOP VAM TOP 3500...". Collapse exact repeated halves.
    parts = cleaned.split()
    if len(parts) % 2 == 0 and parts[:len(parts)//2] == parts[len(parts)//2:]:
        cleaned = " ".join(parts[:len(parts)//2])

    return cleaned or ""


def _extract_connection_values(raw_text: str) -> list[str]:
    """Extract one or multiple connection names from PDF text itself.

    Priority is the exact report fields:
    - Connection Used:
    - Lot(s) Used

    It does not require a hardcoded connection list. It takes the field text and
    removes specs like size/grade/weight only.
    """
    raw = (raw_text or "").replace("\r", "\n")
    found = []

    patterns = [
        r"CONNECTION\s+USED\s*[:\-]?\s*([^\n]+)",
        r"LOT\(S\)\s+USED\s*[:\-]?\s*([^\n]+)",
        r"LOTS?\s+USED\s*[:\-]?\s*([^\n]+)",
    ]

    for pat in patterns:
        for m in re.finditer(pat, raw, flags=re.I):
            line = m.group(1) or ""
            conn = _strip_connection_specs(line)
            if conn and conn not in found:
                found.append(conn)

    # Some PDFs have Manufacturer/Type table instead of a direct connection field.
    # Example: Manufacturer TENARS, Type BLUE -> TENARIS BLUE.
    if not found:
        manu = re.search(r"\bMANUFACTURER\s+([A-Z0-9 ._/-]{2,40})", raw, flags=re.I)
        typ = re.search(r"\bTYPE\s+([A-Z0-9 ._/-]{2,40})", raw, flags=re.I)
        if manu and typ:
            conn = _strip_connection_specs(manu.group(1) + " " + typ.group(1))
            if conn:
                found.append(conn)

    # Fallback from filename/free text only if the explicit fields were missing.
    if not found:
        free = raw.upper().replace("_", " ")
        free = free.replace("TENARS", "TENARIS")
        for phrase in [
            "VAM TOP", "TENARIS BLUE", "TENARS BLUE", "PH-6", "PH6", "PH 6", "PH-7", "PH7", "PH 7",
            "BTC", "LTC", "STC", "EUE", "NUE", "BUTTRESS", "HYDRIL", "XT", "PREMIUM"
        ]:
            if _norm_token(phrase) in _norm_token(free):
                conn = _strip_connection_specs(phrase)
                if conn and conn not in found:
                    found.append(conn)

    return found


def _extract_company_value(raw_text: str, filename: str = "") -> str:
    raw = (raw_text or "").upper()
    compact = _norm_token(raw + " " + (filename or ""))

    # Explicit company fields, but ignore placeholders.
    m = re.search(r"\bCOMPANY\s+([A-Z0-9 ._/-]{2,40})", raw)
    if m:
        val = _clean_field(m.group(1).split("\n")[0]).upper()
        if val and "ENTER COMPANY" not in val:
            if "KHADA" in val or "KHALDA" in val:
                return "KHALDA"
            if "AGIBA" in val:
                return "AGIBA"
            if "SIMETAR" in val or "SIMTAR" in val:
                return "Simetar"

    for name in ["KHALDA", "AGIBA", "SIMETAR"]:
        if name in raw or name in compact:
            return "KHALDA" if name == "KHALDA" else ("Simetar" if name == "SIMETAR" else "AGIBA")
    # Common typo in generated reports.
    if "KHADA" in raw or "KHADA" in compact:
        return "KHALDA"
    return ""


_MONTH_NAMES = {
    1: "01-January", 2: "02-February", 3: "03-March", 4: "04-April",
    5: "05-May", 6: "06-June", 7: "07-July", 8: "08-August",
    9: "09-September", 10: "10-October", 11: "11-November", 12: "12-December"
}

def _month_label(month_num: int | str | None) -> str:
    try:
        m = int(month_num)
    except Exception:
        return datetime.now().strftime("%m-%B")
    return _MONTH_NAMES.get(m, datetime.now().strftime("%m-%B"))

def _extract_job_month(raw_text: str) -> str:
    """Extract the operational month from the PDF.

    Priority:
    1) Start Date Time / End Date Time: YYYY-MM-DD
    2) Report Date fields: DD/MM/YYYY or MM/DD/YYYY
    3) Current month fallback
    """
    raw = raw_text or ""

    # Best source in MTT PDFs.
    m = re.search(r"\b(?:START|END)\s+DATE\s+TIME\s+(\d{4})-(\d{1,2})-\d{1,2}", raw, flags=re.I)
    if m:
        return _month_label(m.group(2))

    # Generic ISO date fallback.
    m = re.search(r"\b(20\d{2})-(\d{1,2})-\d{1,2}\b", raw)
    if m:
        return _month_label(m.group(2))

    # Report Date fallback. If ambiguous, Egyptian reports often use DD/MM/YYYY,
    # but some MTT versions produce M/D/YYYY. Use sane rules.
    m = re.search(r"\bDATE\s+(\d{1,2})/(\d{1,2})/(20\d{2})\b", raw, flags=re.I)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if a > 12 and 1 <= b <= 12:
            return _month_label(b)      # DD/MM/YYYY
        if b > 12 and 1 <= a <= 12:
            return _month_label(a)      # MM/DD/YYYY
        # Ambiguous: prefer first component only when it matches ISO/MTT style,
        # otherwise fallback to second component. Start Date usually resolves this first.
        return _month_label(a)

    return datetime.now().strftime("%m-%B")


def detect_job_file_path(filename: str, content: bytes | None = None) -> dict:
    """Detect Company / Year / Rig / Connection Type from filename + first PDF pages.

    Final TRS filing structure:
        Company → Year → Month → Rig → Connection Type → Files

    Connection Type is now extracted from the PDF text itself, especially from:
        - Connection Used:
        - Lot(s) Used

    If multiple connection fields exist in the same report, they are joined as:
        TENARIS BLUE / VAM TOP
    """
    text = f"{filename or ''}\n"
    if content and file_ext(filename) == "pdf":
        text += extract_pdf_text_sample(content, max_pages=3)

    raw = text.upper()
    compact = _norm_token(text)
    reasons = []
    score = 0

    company = _extract_company_value(text, filename)
    if company:
        reasons.append(f"company={company}")
        score += 25

    y = re.search(r"\b(20\d{2})\b", raw)
    job_year = y.group(1) if y else str(datetime.now().year)
    if job_year:
        reasons.append(f"year={job_year}")
        score += 15

    job_month = _extract_job_month(text)
    if job_month:
        reasons.append(f"month={job_month}")
        score += 10

    rig = ""
    rig_match = re.search(r"\b(EDC|ST|RIG|SINO|HH|IDC)\s*[-_ ]\s*(\d{1,3})\b", raw)
    if not rig_match:
        rig_match = re.search(r"\b(EDC|ST|RIG|SINO|HH|IDC)(\d{1,3})\b", compact)
    if rig_match:
        rig = _normalize_rig(rig_match.group(1), rig_match.group(2))
        reasons.append(f"rig={rig}")
        score += 20

    well_name = ""
    well_match = re.search(r"\bWELL\s+([A-Z0-9][A-Z0-9 ._/-]{1,40})", raw)
    if well_match:
        candidate = well_match.group(1).split("\n")[0]
        candidate = re.split(r"\b(CONTRACTOR|RIG|TECH|FOOTAGE|QUANTITY|TONG|START|END)\b", candidate)[0]
        well_name = _normalize_well(candidate)
    if not well_name and rig:
        pattern = re.escape(rig).replace("\\ ", r"[\s_-]*") + r"\s*[-_ ]+\s*([A-Z0-9][A-Z0-9._ -]{1,35})"
        m = re.search(pattern, raw)
        if m:
            well_name = _normalize_well(m.group(1))
    if well_name:
        reasons.append(f"well={well_name}")
        score += 10

    connections = _extract_connection_values(text)
    connection_name = " / ".join(connections)
    if connection_name:
        reasons.append(f"connection={connection_name}")
        score += 30

    job_name = ""
    jm = re.search(r"\bJOB\s+(JOB#\s*[^\n\r]+)", raw)
    if jm:
        job_name = _clean_field(jm.group(1))
    elif filename:
        job_name = re.sub(r"\.[A-Za-z0-9]+$", "", filename)

    report_date = ""
    dm = re.search(r"\bDATE\s+(\d{1,2}/\d{1,2}/\d{4})\b", raw)
    if dm:
        report_date = dm.group(1)

    company = company or "UNKNOWN COMPANY"
    rig = rig or "UNKNOWN RIG"
    connection_name = connection_name or "UNKNOWN CONNECTION"
    well_name = well_name or "GENERAL"
    detected_path = f"{company}/{job_year}/{job_month}/{rig}/{connection_name}"
    return {
        "company": company,
        "job_year": job_year,
        "job_month": job_month,
        "rig": rig,
        "connection_name": connection_name,
        "well_name": well_name,
        "job_name": job_name,
        "report_date": report_date,
        "detected_path": detected_path,
        "confidence": min(score, 100),
        "needs_review": score < 80,
        "reasons": reasons,
    }


def excel_col_name(n: int) -> str:
    name = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        name = chr(65 + rem) + name
    return name or "A"


def import_equipment_sheet_from_file(equipment_id: int, file_path: str, user_id: int | None = None, max_rows: int = 200, max_cols: int = 30) -> dict:
    """Read an uploaded Excel/CSV equipment file into editable database cells."""
    ext = file_ext(file_path)
    rows: list[list[str]] = []
    sheet_name = "Sheet1"
    try:
        if ext in ["xlsx", "xlsm", "xltx", "xltm"]:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=False)
            ws = wb.active
            sheet_name = ws.title or "Sheet1"
            mr = min(ws.max_row or 0, max_rows)
            mc = min(ws.max_column or 0, max_cols)
            for r in range(1, mr + 1):
                row = []
                for c in range(1, mc + 1):
                    v = ws.cell(r, c).value
                    row.append("" if v is None else str(v))
                rows.append(row)
        elif ext == "csv":
            with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader, start=1):
                    if i > max_rows: break
                    rows.append([str(x) for x in row[:max_cols]])
        else:
            return {"ok": False, "error": "Unsupported sheet file"}
    except Exception as e:
        return {"ok": False, "error": str(e)}

    max_row = len(rows)
    max_col = max((len(r) for r in rows), default=0)
    qx("DELETE FROM trs_equipment_sheet_cells WHERE equipment_id=?", (equipment_id,))
    qx("INSERT OR REPLACE INTO trs_equipment_sheets (equipment_id, sheet_name, max_row, max_col, updated_by, updated_at) VALUES (?,?,?,?,?,datetime('now'))",
       (equipment_id, sheet_name, max_row, max_col, user_id))
    for ri, row in enumerate(rows, start=1):
        for ci, value in enumerate(row, start=1):
            if value != "":
                qx("""
                    INSERT OR REPLACE INTO trs_equipment_sheet_cells (equipment_id,row_index,col_index,value,updated_by,updated_at)
                    VALUES (?,?,?,?,?,datetime('now'))
                """, (equipment_id, ri, ci, value, user_id))
    return {"ok": True, "rows": max_row, "cols": max_col, "sheet_name": sheet_name}


def get_equipment_sheet_matrix(equipment_id: int, min_rows: int = 20, min_cols: int = 8) -> dict:
    meta = q1("SELECT * FROM trs_equipment_sheets WHERE equipment_id=?", (equipment_id,)) or {}
    cells = qa("SELECT row_index, col_index, value FROM trs_equipment_sheet_cells WHERE equipment_id=?", (equipment_id,))
    max_row = max([int(meta.get("max_row") or 0)] + [c["row_index"] for c in cells] + [min_rows])
    max_col = max([int(meta.get("max_col") or 0)] + [c["col_index"] for c in cells] + [min_cols])
    matrix = [["" for _ in range(max_col)] for __ in range(max_row)]
    for c in cells:
        r = c["row_index"] - 1; col = c["col_index"] - 1
        if 0 <= r < max_row and 0 <= col < max_col:
            matrix[r][col] = c.get("value") or ""
    return {"meta": meta, "rows": matrix, "max_row": max_row, "max_col": max_col, "cols": [excel_col_name(i) for i in range(1, max_col+1)]}


def get_equipment_quick_form_view(equipment_id: int) -> dict:
    """Build a simple form/table view from the imported Excel cells.

    It detects the most likely header row from the first 25 rows and turns
    the columns into normal form fields so users can add/edit/delete records
    without editing the full Excel grid.
    """
    meta = q1("SELECT * FROM trs_equipment_sheets WHERE equipment_id=?", (equipment_id,)) or {}
    cells = qa("SELECT row_index, col_index, value FROM trs_equipment_sheet_cells WHERE equipment_id=? ORDER BY row_index, col_index", (equipment_id,))
    if not cells:
        return {"headers": [], "rows": [], "header_row": 1, "next_row": 1, "max_col": 0}

    by_row: dict[int, dict[int, str]] = {}
    max_row = int(meta.get("max_row") or 0)
    max_col = int(meta.get("max_col") or 0)
    for c in cells:
        r = int(c["row_index"]); col = int(c["col_index"]); v = str(c.get("value") or "").strip()
        max_row = max(max_row, r); max_col = max(max_col, col)
        by_row.setdefault(r, {})[col] = v

    # choose header row: first row with at least 2 useful text cells, otherwise row 1
    header_row = 1
    best_score = -1
    for r in range(1, min(max_row, 25) + 1):
        vals = [v for v in by_row.get(r, {}).values() if str(v).strip()]
        score = len(vals)
        joined = " ".join(vals).lower()
        if any(k in joined for k in ["date", "service", "inspection", "description", "done", "remarks", "notes", "status"]):
            score += 3
        if score > best_score and len(vals) >= 2:
            best_score = score
            header_row = r

    headers = []
    used_names = set()
    for col in range(1, max_col + 1):
        name = (by_row.get(header_row, {}).get(col) or "").strip()
        if not name:
            # skip completely empty columns around logos/spacing
            has_data = any((by_row.get(r, {}).get(col) or "").strip() for r in range(header_row + 1, max_row + 1))
            if not has_data:
                continue
            name = f"Column {excel_col_name(col)}"
        name = re.sub(r"\s+", " ", name).strip()[:80]
        base = name
        i = 2
        while name.lower() in used_names:
            name = f"{base} {i}"; i += 1
        used_names.add(name.lower())
        headers.append({"col": col, "name": name})

    rows = []
    for r in range(header_row + 1, max_row + 1):
        values = []
        has_any = False
        for h in headers:
            val = by_row.get(r, {}).get(h["col"], "")
            if str(val).strip():
                has_any = True
            values.append({"col": h["col"], "name": h["name"], "value": val})
        if has_any:
            rows.append({"row_index": r, "values": values})

    return {"headers": headers, "rows": rows[-80:], "header_row": header_row, "next_row": max_row + 1, "max_col": max_col}




def _compact_label(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def get_equipment_maintenance_wizard_view(equipment_id: int) -> dict:
    """Build a guided maintenance-entry form that writes back to the imported Excel columns.

    It reads the real Excel headers from the sheet and maps the requested business fields:
    Sl.no → Maintenance Details → Spare Part Number Used → Spare Part Type
    → Type Of Maintenance → Total Working HRS → Work Order No.
    → Date & Time Finished Maintenance → Maintained By.
    """
    quick = get_equipment_quick_form_view(equipment_id)
    cells = qa("SELECT row_index, col_index, value FROM trs_equipment_sheet_cells WHERE equipment_id=? ORDER BY row_index, col_index", (equipment_id,))
    meta = q1("SELECT * FROM trs_equipment_sheets WHERE equipment_id=?", (equipment_id,)) or {}
    max_row = int(meta.get("max_row") or 0)
    by_row: dict[int, dict[int, str]] = {}
    by_col_text: dict[int, str] = {}
    for c in cells:
        r = int(c["row_index"]); col = int(c["col_index"]); val = str(c.get("value") or "").strip()
        max_row = max(max_row, r)
        by_row.setdefault(r, {})[col] = val
        if r <= 15 and val:
            by_col_text[col] = (by_col_text.get(col, "") + " " + val).strip()

    def find_col(*aliases: str):
        aliases_c = [_compact_label(a) for a in aliases]
        # Prefer detected quick headers first.
        for h in quick.get("headers", []):
            label = _compact_label(h.get("name", ""))
            if any(a and a in label for a in aliases_c):
                return int(h["col"])
        # Then inspect all visible top rows, including merged/grouped headers.
        best = None
        best_score = -1
        for col, txt in by_col_text.items():
            label = _compact_label(txt)
            score = 0
            for a in aliases_c:
                if a and a in label:
                    score = max(score, len(a))
            if score > best_score:
                best_score = score; best = col
        return int(best) if best_score > 0 and best is not None else None

    fields = [
        {"key":"sl_no", "label":"Sl.no", "type":"text", "col":find_col("sl.no", "sl no", "slno", "serial", "s/n")},
        {"key":"maintenance_details", "label":"Maintenance Details", "type":"textarea", "col":find_col("maintenance details", "details", "description", "maintenance description")},
        {"key":"spare_part_number_used", "label":"Spare Part Number Used", "type":"text", "col":find_col("spare part number used", "spare part no", "spare part number", "part number", "part no")},
        {"key":"spare_part_type", "label":"Spare Part Type", "type":"choice", "options":[
            {"key":"local", "label":"Local", "col":find_col("local")},
            {"key":"std", "label":"Std", "col":find_col("std", "standard")},
            {"key":"etc", "label":"Etc", "col":find_col("etc", "other")},
        ]},
        {"key":"type_of_maintenance", "label":"Type Of Maintenance", "type":"choice", "options":[
            {"key":"preventive", "label":"Preventive", "col":find_col("preventive", "preventative", "perventive")},
            {"key":"corrective", "label":"Corrective", "col":find_col("corrective")},
        ]},
        {"key":"total_working_hrs", "label":"Total Working HRS", "type":"number", "col":find_col("total working hrs", "working hrs", "working hours", "total hrs")},
        {"key":"work_order_no", "label":"Work Order No.", "type":"text", "col":find_col("work order no", "work order", "wo no", "w o no")},
        {"key":"date_time_finished", "label":"Date & Time Finished Maintenance", "type":"datetime", "col":find_col("date time finish", "date & time finish", "date and time finish", "finished maintenance", "finish maintenance")},
        {"key":"maintained_by", "label":"Maintained By", "type":"text", "col":find_col("maintained by", "maintenance by", "performed by", "done by", "technician")},
    ]

    # Some Excel templates use merged group headers. In that case only the last
    # sub-column may be detected from OCR/HTML import (for example Corrective),
    # so infer the neighboring columns from the original sheet layout.
    def _get_choice(field_key: str, option_key: str):
        f = next((x for x in fields if x.get("key") == field_key), None)
        if not f:
            return None
        return next((o for o in f.get("options", []) if o.get("key") == option_key), None)

    def _set_choice_col(field_key: str, option_key: str, col):
        if not col or int(col) <= 0:
            return
        opt = _get_choice(field_key, option_key)
        if opt and not opt.get("col"):
            opt["col"] = int(col)

    local = (_get_choice("spare_part_type", "local") or {}).get("col")
    std = (_get_choice("spare_part_type", "std") or {}).get("col")
    etc = (_get_choice("spare_part_type", "etc") or {}).get("col")
    if etc:
        _set_choice_col("spare_part_type", "std", int(etc) - 1)
        _set_choice_col("spare_part_type", "local", int(etc) - 2)
    elif std:
        _set_choice_col("spare_part_type", "local", int(std) - 1)
        _set_choice_col("spare_part_type", "etc", int(std) + 1)
    elif local:
        _set_choice_col("spare_part_type", "std", int(local) + 1)
        _set_choice_col("spare_part_type", "etc", int(local) + 2)

    preventive = (_get_choice("type_of_maintenance", "preventive") or {}).get("col")
    corrective = (_get_choice("type_of_maintenance", "corrective") or {}).get("col")
    if corrective:
        _set_choice_col("type_of_maintenance", "preventive", int(corrective) - 1)
    elif preventive:
        _set_choice_col("type_of_maintenance", "corrective", int(preventive) + 1)

    def cell_value(row: int, col):
        if not col:
            return ""
        return by_row.get(row, {}).get(int(col), "")

    def selected_choice(row: int, field: dict) -> str:
        for opt in field.get("options", []):
            v = str(cell_value(row, opt.get("col"))).strip()
            if v and v not in ["-", "—"]:
                return opt["key"]
        return ""

    # Real data starts after the grouped header row and the sub-header row
    # (for this maintenance template: row 7 group titles, row 8 Local/Std/Etc/Preventive/Corrective).
    data_start_row = int((quick.get("header_row") or 1)) + 2
    header_words = {"local", "std", "etc", "preventive", "corrective", "sl.no", "sl no", "maintenance details", "spare part type", "type of maintenance", "total working hrs", "work order no", "date & time finish", "date time finish", "maintained by"}

    def _is_header_value(v: str) -> bool:
        return re.sub(r"\s+", " ", str(v or "").strip().lower()) in header_words

    rows = []
    for r in quick.get("rows", [])[-120:]:
        row_index = int(r["row_index"])
        if row_index < data_start_row:
            continue
        item = {"row_index": row_index, "values": {}}
        has_any = False
        for f in fields:
            if f["type"] == "choice":
                val = selected_choice(row_index, f)
                display = next((o["label"] for o in f.get("options", []) if o["key"] == val), "")
                item["values"][f["key"]] = {"value": val, "display": display}
                if val: has_any = True
            else:
                val = cell_value(row_index, f.get("col"))
                if _is_header_value(val):
                    val = ""
                item["values"][f["key"]] = {"value": val, "display": val}
                if str(val).strip(): has_any = True
        if has_any:
            rows.append(item)

    mapped_cols = []
    for f in fields:
        if f["type"] == "choice":
            mapped_cols.extend([o.get("col") for o in f.get("options", []) if o.get("col")])
        elif f.get("col"):
            mapped_cols.append(f.get("col"))

    # Pick the first empty maintenance row in the actual Excel area instead of
    # jumping to the end of the imported sheet. This makes the next entry write
    # to row 9, then row 10, etc. when the template has empty prepared rows.
    mapped_cols_i = [int(c) for c in mapped_cols if c]
    first_empty = None
    scan_limit = max(max_row + 2, data_start_row + 50)
    for rr in range(data_start_row, scan_limit + 1):
        if not any(str(by_row.get(rr, {}).get(c, "")).strip() for c in mapped_cols_i):
            first_empty = rr
            break
    if first_empty is None:
        first_empty = max_row + 1

    return {
        "fields": fields,
        "rows": rows[-80:],
        "next_row": first_empty,
        "header_row": quick.get("header_row") or 1,
        "data_start_row": data_start_row,
        "max_col": quick.get("max_col") or 1,
        "ready": bool(mapped_cols),
    }


def apply_equipment_maintenance_wizard_form(equipment_id: int, row_index: int, form, user_id: int | None):
    wizard = get_equipment_maintenance_wizard_view(equipment_id)
    for f in wizard.get("fields", []):
        if f.get("type") == "choice":
            selected = str(form.get(f"wizard_{f['key']}") or "").strip()
            for opt in f.get("options", []):
                col = opt.get("col")
                if not col:
                    continue
                upsert_equipment_sheet_cell(equipment_id, row_index, int(col), "✓" if selected == opt.get("key") else "", user_id)
        else:
            col = f.get("col")
            if not col:
                continue
            upsert_equipment_sheet_cell(equipment_id, row_index, int(col), str(form.get(f"wizard_{f['key']}") or ""), user_id)
    return wizard


def upsert_equipment_sheet_cell(equipment_id: int, row: int, col: int, value: str, user_id: int | None):
    value = (value or "").strip()
    if value == "":
        qx("DELETE FROM trs_equipment_sheet_cells WHERE equipment_id=? AND row_index=? AND col_index=?", (equipment_id, row, col))
    else:
        qx("""
            INSERT INTO trs_equipment_sheet_cells (equipment_id,row_index,col_index,value,updated_by,updated_at)
            VALUES (?,?,?,?,?,datetime('now'))
            ON CONFLICT(equipment_id,row_index,col_index) DO UPDATE SET
              value=excluded.value, updated_by=excluded.updated_by, updated_at=datetime('now')
        """, (equipment_id, row, col, value, user_id))
    try:
        write_equipment_cell_to_source(equipment_id, row, col, value)
    except Exception as e:
        logger.debug(f"Source Excel quick-form save skipped: {e}")


# ══════════════════════════════════════════════════════════════
#  TRS PLATFORM MODULES — Job Files + Equipment Maintenance
# ══════════════════════════════════════════════════════════════

@app.get("/job-files", response_class=HTMLResponse)
async def job_files_page(request: Request, company: str = "", job_year: str = "", job_month: str = "", rig: str = "", connection_name: str = "", q: str = ""):
    user, redir = require_user(request)
    if redir: return redir

    # Current file filter. Well is intentionally not a navigation level anymore.
    where, params = [], []
    for col, val in [("company", company), ("job_year", job_year), ("job_month", job_month), ("rig", rig), ("connection_name", connection_name)]:
        if val:
            where.append(f"{col}=?")
            params.append(val)
    if q:
        where.append("(original_name LIKE ? OR notes LIKE ? OR job_name LIKE ? OR COALESCE(well_name,'') LIKE ?)")
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    files = qa(f"""
        SELECT f.*, u.username, u.full_name
        FROM trs_job_files f
        LEFT JOIN users u ON u.id=f.uploaded_by
        {where_sql}
        ORDER BY f.company, f.job_year DESC, f.job_month, f.rig, f.connection_name, f.created_at DESC
        LIMIT 300
    """, tuple(params))

    # Flat connection-level tree: Company / Year / Rig / Connection.
    tree = qa("""
        SELECT company, job_year, job_month, rig, connection_name, SUM(file_count) AS file_count, MAX(last_update) AS last_update
        FROM (
            SELECT company, job_year, job_month, rig, connection_name, COUNT(*) file_count, MAX(created_at) last_update
            FROM trs_job_files
            GROUP BY company, job_year, job_month, rig, connection_name
            UNION ALL
            SELECT company, job_year, job_month, rig, connection_name, 0 file_count, created_at last_update
            FROM trs_job_folders
        ) x
        GROUP BY company, job_year, job_month, rig, connection_name
        ORDER BY company, job_year DESC, job_month, rig, connection_name
    """)

    # Browser lists to keep the UI separated instead of showing everything mixed.
    companies = qa("""
        SELECT company, COUNT(*) c
        FROM (
            SELECT company FROM trs_job_files
            UNION ALL
            SELECT company FROM trs_job_folders
        )
        GROUP BY company
        ORDER BY company
    """)

    years = qa("""
        SELECT job_year, COUNT(*) c
        FROM (
            SELECT company, job_year FROM trs_job_files
            UNION ALL
            SELECT company, job_year FROM trs_job_folders
        )
        WHERE (?='' OR company=?)
        GROUP BY job_year
        ORDER BY job_year DESC
    """, (company, company))

    months = qa("""
        SELECT job_month, COUNT(*) c
        FROM (
            SELECT company, job_year, job_month FROM trs_job_files
            UNION ALL
            SELECT company, job_year, job_month FROM trs_job_folders
        )
        WHERE (?='' OR company=?) AND (?='' OR job_year=?)
        GROUP BY job_month
        ORDER BY job_month
    """, (company, company, job_year, job_year))

    rigs = qa("""
        SELECT rig, COUNT(*) c
        FROM (
            SELECT company, job_year, job_month, rig FROM trs_job_files
            UNION ALL
            SELECT company, job_year, job_month, rig FROM trs_job_folders
        )
        WHERE (?='' OR company=?) AND (?='' OR job_year=?) AND (?='' OR job_month=?)
        GROUP BY rig
        ORDER BY rig
    """, (company, company, job_year, job_year, job_month, job_month))

    connections = qa("""
        SELECT connection_name, COUNT(*) c
        FROM (
            SELECT company, job_year, job_month, rig, connection_name FROM trs_job_files
            UNION ALL
            SELECT company, job_year, job_month, rig, connection_name FROM trs_job_folders
        )
        WHERE (?='' OR company=?) AND (?='' OR job_year=?) AND (?='' OR job_month=?) AND (?='' OR rig=?)
        GROUP BY connection_name
        ORDER BY connection_name
    """, (company, company, job_year, job_year, job_month, job_month, rig, rig))

    stats = {
        "files": q1("SELECT COUNT(*) c FROM trs_job_files")["c"],
        "companies": q1("SELECT COUNT(DISTINCT company) c FROM (SELECT company FROM trs_job_files UNION SELECT company FROM trs_job_folders)")["c"],
        "connections": q1("SELECT COUNT(*) c FROM (SELECT company, job_year, job_month, rig, connection_name FROM trs_job_files UNION SELECT company, job_year, job_month, rig, connection_name FROM trs_job_folders)")["c"],
    }

    return render(request, "job_files.html", {
        "active": "job_files", "files": files, "tree": tree, "companies": companies,
        "years": years, "months": months, "rigs": rigs, "connections": connections,
        "stats": stats,
        "filters": {"company": company, "job_year": job_year, "job_month": job_month, "rig": rig, "connection_name": connection_name, "q": q}
    })




@app.post("/job-files/folders/create")
async def job_files_create_folder(request: Request,
                                  company: str = Form(...), job_year: str = Form(...), job_month: str = Form(""),
                                  rig: str = Form(...), connection_name: str = Form(...), notes: str = Form("")):
    user, redir = require_user(request)
    if redir: return redir
    company, job_year, job_month, rig, connection_name = company.strip(), job_year.strip(), (job_month.strip() or datetime.now().strftime("%m-%B")), rig.strip(), connection_name.strip()
    well_name = "GENERAL"
    qx("""
        INSERT OR IGNORE INTO trs_job_folders (company, job_year, job_month, rig, connection_name, well_name, notes, created_by)
        VALUES (?,?,?,?,?,?,?,?)
    """, (company, job_year, job_month, rig, connection_name, well_name, notes.strip(), user["id"]))
    folder_path = f"{company}/{job_year}/{job_month}/{rig}/{connection_name}"
    os.makedirs(os.path.join(str(UPLOADS_DIR), "job_files", safe_name(company), safe_name(job_year), safe_name(job_month), safe_name(rig), safe_name(connection_name)), exist_ok=True)
    log_activity("job_folder_create", user["id"], "trs_job_folder", None, f"Created folder {folder_path}")
    await ws_manager.broadcast({"type":"job_folder_create", "path": folder_path}, "job_files")
    return RedirectResponse(f"/job-files?company={company}&job_year={job_year}&job_month={job_month}&rig={rig}&connection_name={connection_name}", status_code=303)


@app.post("/job-files/import-zip")
async def job_files_import_zip(request: Request, file: UploadFile = File(...)):
    """Import a ZIP of job files. Each file is auto-detected and stored under the suggested tree.
    Useful when the user receives a batch of PDFs/Excels and wants TRS to organize them.
    """
    user, redir = require_user(request)
    if redir: return redir
    content = await file.read()
    if not zipfile.is_zipfile(io.BytesIO(content)):
        return JSONResponse({"error":"Please upload a .zip file."}, status_code=400)
    imported = 0
    review = 0
    skipped = 0
    allowed = {"pdf","xlsx","xls","csv","png","jpg","jpeg"}
    with zipfile.ZipFile(io.BytesIO(content)) as z:
        for info in z.infolist():
            if info.is_dir():
                continue
            original = os.path.basename(info.filename)
            ext = file_ext(original)
            if not original or original.startswith("~$") or ext not in allowed:
                skipped += 1
                continue
            data = z.read(info)
            detected = detect_job_file_path(original, data)
            company, job_year, job_month, rig, connection_name = detected["company"], detected["job_year"], detected.get("job_month") or datetime.now().strftime("%m-%B"), detected["rig"], detected["connection_name"]
            well_name = detected.get("well_name") or "GENERAL"  # metadata only
            folder = os.path.join(str(UPLOADS_DIR), "job_files", safe_name(company), safe_name(job_year), safe_name(job_month), safe_name(rig), safe_name(connection_name))
            os.makedirs(folder, exist_ok=True)
            stored = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}_{safe_name(original)}"
            path = os.path.join(folder, stored)
            with open(path, "wb") as out:
                out.write(data)
            fid = qx("""
                INSERT INTO trs_job_files (company, job_year, job_month, rig, connection_name, well_name, job_name, report_date, original_name, stored_name, file_path, file_type, file_size, notes, detection_confidence, detected_path, uploaded_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (company, job_year, job_month, rig, connection_name, well_name, detected.get("job_name") or "", detected.get("report_date") or "", original, stored, path, ext, len(data), f"Batch imported from {file.filename}", detected["confidence"], detected.get("detected_path") or f"{company}/{job_year}/{job_month}/{rig}/{connection_name}", user["id"]))
            qx("""INSERT OR IGNORE INTO trs_job_folders (company, job_year, job_month, rig, connection_name, well_name, notes, created_by) VALUES (?,?,?,?,?,?,?,?)""", (company, job_year, job_month, rig, connection_name, well_name, "Auto-created by ZIP import", user["id"]))
            imported += 1
            if detected.get("needs_review"):
                review += 1
    log_activity("job_zip_import", user["id"], "trs_job_file", None, f"Imported {imported} job files from {file.filename}; {review} need review; {skipped} skipped")
    await ws_manager.broadcast({"type":"job_zip_import", "imported": imported, "review": review, "skipped": skipped}, "job_files")
    return RedirectResponse("/job-files", status_code=303)


@app.post("/job-files/auto-detect")
async def job_files_auto_detect(request: Request, file: UploadFile = File(...)):
    user, redir = require_user(request)
    if redir: return redir
    content = await file.read()
    return JSONResponse(detect_job_file_path(file.filename or "upload", content))


@app.post("/job-files/upload")
async def job_files_upload(request: Request,
                           company: str = Form(""), job_year: str = Form(""), job_month: str = Form(""), rig: str = Form(""),
                           connection_name: str = Form(""), notes: str = Form(""),
                           file: UploadFile = File(...)):
    user, redir = require_user(request)
    if redir: return redir

    original = file.filename or "upload.bin"
    content = await file.read()
    detected = detect_job_file_path(original, content)
    if not (company and job_year and job_month and rig and connection_name):
        company = company or detected["company"]
        job_year = job_year or detected["job_year"]
        job_month = job_month or detected.get("job_month") or datetime.now().strftime("%m-%B")
        rig = rig or detected["rig"]
        connection_name = connection_name or detected["connection_name"]
    company = (company or "UNKNOWN COMPANY").strip()
    job_year = (job_year or str(datetime.now().year)).strip()
    job_month = (job_month or datetime.now().strftime("%m-%B")).strip()
    rig = (rig or "UNKNOWN RIG").strip()
    connection_name = (connection_name or "UNKNOWN CONNECTION").strip()
    well_name = detected.get("well_name") or "GENERAL"  # metadata only, not folder level

    folder = os.path.join(str(UPLOADS_DIR), "job_files", safe_name(company), safe_name(job_year), safe_name(job_month), safe_name(rig), safe_name(connection_name))
    os.makedirs(folder, exist_ok=True)
    stored = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}_{safe_name(original)}"
    path = os.path.join(folder, stored)
    with open(path, "wb") as f:
        f.write(content)

    detected_path = detected.get("detected_path") or f"{company}/{job_year}/{job_month}/{rig}/{connection_name}"
    fid = qx("""
        INSERT INTO trs_job_files (company, job_year, job_month, rig, connection_name, well_name, job_name, report_date, original_name, stored_name, file_path, file_type, file_size, notes, detection_confidence, detected_path, confirmed_by, confirmed_at, uploaded_by)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (company.strip(), job_year.strip(), job_month.strip(), rig.strip(), connection_name.strip(), (well_name or "GENERAL").strip(), detected.get("job_name") or "", detected.get("report_date") or "", original, stored, path, file_ext(original), len(content), notes.strip(), int(detected.get("confidence") or 0), detected_path, user["id"], datetime.now().isoformat(timespec="seconds"), user["id"]))
    log_activity("job_file_upload", user["id"], "trs_job_file", fid, f"Uploaded {original} to {company}/{job_year}/{job_month}/{rig}/{connection_name}")
    await ws_manager.broadcast({"type":"job_file_upload", "file_id": fid, "path": f"{company}/{job_year}/{job_month}/{rig}/{connection_name}"}, "job_files")
    return RedirectResponse(f"/job-files?company={company}&job_year={job_year}&job_month={job_month}&rig={rig}&connection_name={connection_name}", status_code=303)



@app.get("/job-files/{file_id}", response_class=HTMLResponse)
async def job_file_detail(request: Request, file_id: int):
    user, redir = require_user(request)
    if redir: return redir
    f = q1("""
        SELECT f.*, u.username, u.full_name
        FROM trs_job_files f LEFT JOIN users u ON u.id=f.uploaded_by
        WHERE f.id=?
    """, (file_id,))
    if not f:
        return RedirectResponse("/job-files", status_code=303)
    activity = qa("""
        SELECT a.*, u.username, u.full_name
        FROM activity_log a LEFT JOIN users u ON u.id=a.user_id
        WHERE a.entity_type='trs_job_file' AND a.entity_id=?
        ORDER BY a.created_at DESC LIMIT 20
    """, (file_id,))
    return render(request, "job_file_detail.html", {"active":"job_files", "f": f, "activity": activity})


@app.post("/job-files/{file_id}/move")
async def job_files_move(request: Request, file_id: int,
                         company: str = Form(...), job_year: str = Form(...), job_month: str = Form(""), rig: str = Form(...), connection_name: str = Form(...), notes: str = Form("")):
    user, redir = require_user(request)
    if redir: return redir
    f = q1("SELECT * FROM trs_job_files WHERE id=?", (file_id,))
    if not f:
        return RedirectResponse("/job-files", status_code=303)
    old_path = f"{f['company']}/{f['job_year']}/{f['job_month'] or ''}/{f['rig']}/{f['connection_name']}".replace('//','/')
    job_month = job_month or f["job_month"] or datetime.now().strftime("%m-%B")
    new_path = f"{company}/{job_year}/{job_month}/{rig}/{connection_name}"
    # Preserve the detected well as metadata if it already exists, but do not use it as a folder level.
    well_name = f.get("well_name") or "GENERAL"
    qx("""
        UPDATE trs_job_files SET company=?, job_year=?, job_month=?, rig=?, connection_name=?, well_name=?, notes=?, confirmed_by=?, confirmed_at=datetime('now') WHERE id=?
    """, (company.strip(), job_year.strip(), job_month.strip(), rig.strip(), connection_name.strip(), well_name, notes.strip(), user["id"], file_id))
    log_activity("job_file_move", user["id"], "trs_job_file", file_id, f"Moved file path: {old_path} → {new_path}")
    await ws_manager.broadcast({"type":"job_file_move", "file_id": file_id, "path": new_path}, "job_files")
    return RedirectResponse(f"/job-files/{file_id}", status_code=303)


@app.get("/job-files/{file_id}/download")
async def job_files_download(request: Request, file_id: int):
    user, redir = require_user(request)
    if redir: return redir
    f = q1("SELECT * FROM trs_job_files WHERE id=?", (file_id,))
    if not f or not os.path.exists(f["file_path"]):
        return JSONResponse({"error":"File not found"}, status_code=404)
    return FileResponse(f["file_path"], filename=f["original_name"])


@app.get("/job-files/{file_id}/preview")
async def job_files_preview(request: Request, file_id: int):
    user, redir = require_user(request)
    if redir: return redir
    f = q1("SELECT * FROM trs_job_files WHERE id=?", (file_id,))
    if not f or not os.path.exists(f["file_path"]):
        return JSONResponse({"error":"File not found"}, status_code=404)
    media_type = "application/pdf" if f.get("file_type") == "pdf" else None
    return FileResponse(f["file_path"], media_type=media_type, filename=f["original_name"])


@app.post("/job-files/{file_id}/delete")
async def job_files_delete(request: Request, file_id: int):
    user, redir = require_role(request, ["Admin", "Manager"])
    if redir: return redir
    f = q1("SELECT * FROM trs_job_files WHERE id=?", (file_id,))
    if f:
        try:
            if os.path.exists(f["file_path"]): os.remove(f["file_path"])
        except Exception as e: logger.debug(f"Ignored non-critical error: {e}")
        qx("DELETE FROM trs_job_files WHERE id=?", (file_id,))
        log_activity("job_file_delete", user["id"], "trs_job_file", file_id, f"Deleted {f['original_name']}")
        await ws_manager.broadcast({"type":"job_file_delete", "file_id": file_id}, "job_files")
    return RedirectResponse("/job-files", status_code=303)



@app.post("/job-files/folders/delete")
async def job_files_folder_delete(request: Request, company: str = Form(...), job_year: str = Form(...), job_month: str = Form(""), rig: str = Form(...), connection_name: str = Form(...)):
    user, redir = require_role(request, ["Admin", "Manager"])
    if redir: return redir
    # Only delete empty connection-level folder records; files remain protected.
    files_count = q1("SELECT COUNT(*) c FROM trs_job_files WHERE company=? AND job_year=? AND job_month=? AND rig=? AND connection_name=?", (company, job_year, job_month, rig, connection_name))["c"]
    if files_count == 0:
        qx("DELETE FROM trs_job_folders WHERE company=? AND job_year=? AND job_month=? AND rig=? AND connection_name=?", (company, job_year, job_month, rig, connection_name))
        log_activity("job_folder_delete", user["id"], "trs_job_folder", None, f"Deleted empty folder {company}/{job_year}/{job_month}/{rig}/{connection_name}")
        await ws_manager.broadcast({"type":"job_folder_delete", "path": f"{company}/{job_year}/{job_month}/{rig}/{connection_name}"}, "job_files")
    return RedirectResponse("/job-files", status_code=303)


@app.get("/equipment-maintenance", response_class=HTMLResponse)
async def equipment_page(request: Request, status: str = "", q: str = ""):
    user, redir = require_user(request)
    if redir: return redir
    where, params = [], []
    if status:
        where.append("status=?"); params.append(status)
    if q:
        where.append("(name LIKE ? OR category LIKE ? OR serial_no LIKE ? OR location LIKE ?)")
        params.extend([f"%{q}%"]*4)
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    equipment = qa(f"""
        SELECT e.*,
               CASE
                 WHEN e.next_service_date IS NOT NULL AND date(e.next_service_date) < date('now') THEN 'OVERDUE'
                 WHEN e.next_service_date IS NOT NULL AND date(e.next_service_date) <= date('now','+30 day') THEN 'DUE_SOON'
                 ELSE 'OK'
               END AS service_state,
               (SELECT COUNT(*) FROM trs_equipment_logs l WHERE l.equipment_id=e.id) AS log_count
        FROM trs_equipment e
        {where_sql}
        ORDER BY e.category COLLATE NOCASE, service_state DESC, e.next_service_date, e.name
    """, tuple(params))
    stats = {
        "total": q1("SELECT COUNT(*) c FROM trs_equipment")["c"],
        "available": q1("SELECT COUNT(*) c FROM trs_equipment WHERE status='Available'")["c"],
        "overdue": q1("SELECT COUNT(*) c FROM trs_equipment WHERE next_service_date IS NOT NULL AND date(next_service_date) < date('now')")["c"],
        "due_soon": q1("SELECT COUNT(*) c FROM trs_equipment WHERE next_service_date IS NOT NULL AND date(next_service_date) BETWEEN date('now') AND date('now','+30 day')")["c"],
    }
    category_stats = qa("""
        SELECT category, COUNT(*) AS total,
               SUM(CASE WHEN status='Available' THEN 1 ELSE 0 END) AS available,
               SUM(CASE WHEN status='Maintenance' THEN 1 ELSE 0 END) AS maintenance,
               SUM(CASE WHEN status='Out of Service' THEN 1 ELSE 0 END) AS out_of_service
        FROM trs_equipment
        GROUP BY category
        ORDER BY category COLLATE NOCASE
    """)
    return render(request, "equipment.html", {"active":"equipment", "equipment": equipment, "stats": stats, "category_stats": category_stats, "filters": {"status": status, "q": q}})


@app.post("/equipment-maintenance/create")
async def equipment_create(request: Request,
                           name: str = Form(...), category: str = Form(""), serial_no: str = Form(""),
                           manufacture_date: str = Form(""), last_service_date: str = Form(""), next_service_date: str = Form(""),
                           status: str = Form("Available"), location: str = Form(""), notes: str = Form("")):
    user, redir = require_user(request)
    if redir: return redir
    eid = qx("""
        INSERT INTO trs_equipment (name, category, serial_no, manufacture_date, last_service_date, next_service_date, status, location, notes, created_by)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (name.strip(), category.strip(), serial_no.strip(), manufacture_date or None, last_service_date or None, next_service_date or None, status, location.strip(), notes.strip(), user["id"]))
    log_activity("equipment_create", user["id"], "trs_equipment", eid, f"Created equipment {name}")
    await ws_manager.broadcast({"type":"equipment_create", "equipment_id": eid}, "equipment")
    return RedirectResponse("/equipment-maintenance", status_code=303)


@app.get("/equipment-maintenance/{equipment_id}", response_class=HTMLResponse)
async def equipment_detail(request: Request, equipment_id: int):
    user, redir = require_user(request)
    if redir: return redir
    eq = q1("SELECT * FROM trs_equipment WHERE id=?", (equipment_id,))
    if not eq:
        return RedirectResponse("/equipment-maintenance", status_code=303)
    logs = qa("""
        SELECT l.*, u.username, u.full_name
        FROM trs_equipment_logs l LEFT JOIN users u ON u.id=l.created_by
        WHERE l.equipment_id=? ORDER BY l.service_date DESC, l.created_at DESC
    """, (equipment_id,))
    sheet = get_equipment_sheet_matrix(equipment_id)
    sheet["html"] = render_styled_equipment_sheet_html(equipment_id, eq)
    quick_form = get_equipment_quick_form_view(equipment_id)
    maintenance_wizard = get_equipment_maintenance_wizard_view(equipment_id)
    return render(request, "equipment_detail.html", {"active":"equipment", "eq": eq, "logs": logs, "sheet": sheet, "quick_form": quick_form, "maintenance_wizard": maintenance_wizard})


@app.post("/equipment-maintenance/{equipment_id}/update")
async def equipment_update(request: Request, equipment_id: int,
                           name: str = Form(...), category: str = Form(""), serial_no: str = Form(""),
                           manufacture_date: str = Form(""), last_service_date: str = Form(""), next_service_date: str = Form(""),
                           status: str = Form("Available"), location: str = Form(""), notes: str = Form("")):
    user, redir = require_user(request)
    if redir: return redir
    qx("""
        UPDATE trs_equipment SET name=?, category=?, serial_no=?, manufacture_date=?, last_service_date=?, next_service_date=?, status=?, location=?, notes=?, updated_at=datetime('now')
        WHERE id=?
    """, (name.strip(), category.strip(), serial_no.strip(), manufacture_date or None, last_service_date or None, next_service_date or None, status, location.strip(), notes.strip(), equipment_id))
    log_activity("equipment_update", user["id"], "trs_equipment", equipment_id, f"Updated equipment {name}")
    await ws_manager.broadcast({"type":"equipment_update", "equipment_id": equipment_id}, "equipment")
    return RedirectResponse(f"/equipment-maintenance/{equipment_id}", status_code=303)


@app.post("/equipment-maintenance/{equipment_id}/logs/create")
async def equipment_log_create(request: Request, equipment_id: int,
                               service_date: str = Form(...), service_type: str = Form("Maintenance"),
                               performed_by: str = Form(""), next_due_date: str = Form(""), notes: str = Form("")):
    user, redir = require_user(request)
    if redir: return redir
    lid = qx("""
        INSERT INTO trs_equipment_logs (equipment_id, service_date, service_type, performed_by, next_due_date, notes, created_by)
        VALUES (?,?,?,?,?,?,?)
    """, (equipment_id, service_date, service_type.strip(), performed_by.strip(), next_due_date or None, notes.strip(), user["id"]))
    qx("UPDATE trs_equipment SET last_service_date=?, next_service_date=COALESCE(?, next_service_date), updated_at=datetime('now') WHERE id=?", (service_date, next_due_date or None, equipment_id))
    log_activity("equipment_log_create", user["id"], "trs_equipment_log", lid, f"Added maintenance log")
    await ws_manager.broadcast({"type":"equipment_log_create", "equipment_id": equipment_id}, "equipment")
    return RedirectResponse(f"/equipment-maintenance/{equipment_id}", status_code=303)


@app.post("/equipment-maintenance/import-folder")
async def equipment_import_folder(request: Request,
                                  category: str = Form("HPU"),
                                  location: str = Form(""),
                                  file: UploadFile = File(...)):
    """Import a zipped equipment folder, e.g. HPU/111617.xlsx.
    This creates one equipment record per Excel file using the file name as serial/equipment ID.
    """
    user, redir = require_user(request)
    if redir: return redir
    content = await file.read()
    imported = 0
    skipped = 0
    base = os.path.join(str(UPLOADS_DIR), "equipment_imports", safe_name(category), datetime.now().strftime("%Y%m%d_%H%M%S"))
    os.makedirs(base, exist_ok=True)
    if not zipfile.is_zipfile(io.BytesIO(content)):
        return JSONResponse({"error":"Please upload a .zip containing the equipment Excel files."}, status_code=400)
    with zipfile.ZipFile(io.BytesIO(content)) as z:
        for info in z.infolist():
            if info.is_dir():
                continue
            name = os.path.basename(info.filename)
            ext = file_ext(name)
            if ext not in ["xlsx", "xls", "csv"] or name.startswith("~$"):
                skipped += 1
                continue
            equip_id = os.path.splitext(name)[0].strip()
            if not equip_id:
                skipped += 1
                continue
            existing = q1("SELECT id FROM trs_equipment WHERE category=? AND serial_no=?", (category.strip(), equip_id))
            data = z.read(info)
            stored = f"{safe_name(equip_id)}_{uuid.uuid4().hex[:6]}.{ext}"
            path = os.path.join(base, stored)
            with open(path, "wb") as out:
                out.write(data)
            note = f"Imported from folder file: {info.filename}; stored at {path}"
            if existing:
                qx("UPDATE trs_equipment SET notes=?, source_file_path=?, source_original_name=?, updated_at=datetime('now') WHERE id=?", (note, path, name, existing["id"]))
                import_equipment_sheet_from_file(existing["id"], path, user["id"])
                skipped += 1
            else:
                new_eid = qx("""
                    INSERT INTO trs_equipment (name, category, serial_no, status, location, notes, source_file_path, source_original_name, created_by)
                    VALUES (?,?,?,?,?,?,?,?,?)
                """, (f"{category.strip()} {equip_id}", category.strip(), equip_id, "Available", location.strip(), note, path, name, user["id"]))
                import_equipment_sheet_from_file(new_eid, path, user["id"])
                imported += 1
    log_activity("equipment_import", user["id"], "trs_equipment", None, f"Imported {imported} equipment from {file.filename}")
    await ws_manager.broadcast({"type":"equipment_import", "imported": imported, "skipped": skipped}, "equipment")
    return RedirectResponse(f"/equipment-maintenance?q={category}", status_code=303)






@app.post("/equipment-maintenance/{equipment_id}/delete")
async def equipment_delete(request: Request, equipment_id: int):
    user, redir = require_role(request, ["Admin", "Manager"])
    if redir: return redir
    eq = q1("SELECT * FROM trs_equipment WHERE id=?", (equipment_id,))
    if eq:
        try:
            src = resolve_equipment_source_path(eq)
            # Keep source files by default for audit/demo safety. Deleting DB record removes the asset from UI.
        except Exception:
            pass
        qx("DELETE FROM trs_equipment WHERE id=?", (equipment_id,))
        log_activity("equipment_delete", user["id"], "trs_equipment", equipment_id, f"Deleted equipment {eq.get('name') or equipment_id}")
        await ws_manager.broadcast({"type":"equipment_delete", "equipment_id": equipment_id}, "equipment")
    return RedirectResponse("/equipment-maintenance", status_code=303)


@app.post("/equipment-maintenance/{equipment_id}/quick-row/create")
async def equipment_quick_row_create(request: Request, equipment_id: int):
    user, redir = require_user(request)
    if redir: return redir
    eq = q1("SELECT * FROM trs_equipment WHERE id=?", (equipment_id,))
    if not eq:
        return RedirectResponse("/equipment-maintenance", status_code=303)
    form = await request.form()
    quick = get_equipment_quick_form_view(equipment_id)
    if str(form.get("wizard_mode") or "") == "1":
        wizard_view = get_equipment_maintenance_wizard_view(equipment_id)
        new_row = int(form.get("row_index") or wizard_view.get("next_row") or quick.get("next_row") or 1)
        if new_row < int(wizard_view.get("data_start_row") or (quick.get("header_row") or 1)):
            new_row = int(wizard_view.get("next_row") or new_row)
        wiz = apply_equipment_maintenance_wizard_form(equipment_id, new_row, form, user["id"])
        max_col = int(wiz.get("max_col") or quick.get("max_col") or 1)
    else:
        new_row = int(form.get("row_index") or quick.get("next_row") or 1)
        if new_row <= int(quick.get("header_row") or 1):
            new_row = int(quick.get("next_row") or 1)
        for h in quick.get("headers", []):
            col = int(h["col"])
            upsert_equipment_sheet_cell(equipment_id, new_row, col, str(form.get(f"col_{col}") or ""), user["id"])
        max_col = int(quick.get("max_col") or 1)
    qx("INSERT OR REPLACE INTO trs_equipment_sheets (equipment_id,sheet_name,max_row,max_col,updated_by,updated_at) VALUES (?,COALESCE((SELECT sheet_name FROM trs_equipment_sheets WHERE equipment_id=?),'Live Sheet'),?,?,?,datetime('now'))",
       (equipment_id, equipment_id, max(new_row, int(quick.get("next_row") or new_row)), max_col, user["id"]))
    log_activity("equipment_quick_row_create", user["id"], "trs_equipment", equipment_id, f"Added simple entry row {new_row}")
    await ws_manager.broadcast({"type":"equipment_sheet_row_add", "equipment_id": equipment_id, "by": user.get("full_name") or user.get("username")}, "equipment")
    return RedirectResponse(f"/equipment-maintenance/{equipment_id}?saved=1", status_code=303)


@app.post("/equipment-maintenance/{equipment_id}/quick-row/{row_index}/update")
async def equipment_quick_row_update(request: Request, equipment_id: int, row_index: int):
    user, redir = require_user(request)
    if redir: return redir
    eq = q1("SELECT * FROM trs_equipment WHERE id=?", (equipment_id,))
    if not eq:
        return RedirectResponse("/equipment-maintenance", status_code=303)
    form = await request.form()
    quick = get_equipment_quick_form_view(equipment_id)
    if str(form.get("wizard_mode") or "") == "1":
        apply_equipment_maintenance_wizard_form(equipment_id, row_index, form, user["id"])
    else:
        for h in quick.get("headers", []):
            col = int(h["col"])
            upsert_equipment_sheet_cell(equipment_id, row_index, col, str(form.get(f"col_{col}") or ""), user["id"])
    log_activity("equipment_quick_row_update", user["id"], "trs_equipment", equipment_id, f"Updated simple entry row {row_index}")
    await ws_manager.broadcast({"type":"equipment_sheet_update", "equipment_id": equipment_id, "by": user.get("full_name") or user.get("username")}, "equipment")
    return RedirectResponse(f"/equipment-maintenance/{equipment_id}?updated=1", status_code=303)


@app.post("/equipment-maintenance/{equipment_id}/quick-row/{row_index}/delete")
async def equipment_quick_row_delete(request: Request, equipment_id: int, row_index: int):
    user, redir = require_user(request)
    if redir: return redir
    # Clear only the mapped maintenance-entry columns in DB and in the original Excel,
    # so the visual workbook stays synchronized with Add/Edit/Delete actions.
    try:
        wiz = get_equipment_maintenance_wizard_view(equipment_id)
        cols = []
        for f in wiz.get("fields", []):
            if f.get("type") == "choice":
                cols.extend([o.get("col") for o in f.get("options", []) if o.get("col")])
            elif f.get("col"):
                cols.append(f.get("col"))
        cols = sorted({int(c) for c in cols if c})
        for col in cols:
            upsert_equipment_sheet_cell(equipment_id, row_index, col, "", user["id"])
    except Exception as e:
        logger.debug(f"Wizard delete source clear skipped: {e}")
        qx("DELETE FROM trs_equipment_sheet_cells WHERE equipment_id=? AND row_index=?", (equipment_id, row_index))
    log_activity("equipment_quick_row_delete", user["id"], "trs_equipment", equipment_id, f"Deleted simple entry row {row_index}")
    await ws_manager.broadcast({"type":"equipment_sheet_update", "equipment_id": equipment_id, "by": user.get("full_name") or user.get("username")}, "equipment")
    return RedirectResponse(f"/equipment-maintenance/{equipment_id}?deleted=1", status_code=303)


@app.get("/equipment-maintenance/{equipment_id}/sheet.json")
async def equipment_sheet_json(request: Request, equipment_id: int):
    user, redir = require_user(request)
    if redir: return redir
    eq = q1("SELECT id FROM trs_equipment WHERE id=?", (equipment_id,))
    if not eq:
        return JSONResponse({"error":"Equipment not found"}, status_code=404)
    return JSONResponse(get_equipment_sheet_matrix(equipment_id))


@app.post("/equipment-maintenance/{equipment_id}/sheet/cell")
async def equipment_sheet_cell_update(request: Request, equipment_id: int):
    user, redir = require_user(request)
    if redir: return redir
    payload = await request.json()
    row = int(payload.get("row", 0)); col = int(payload.get("col", 0)); value = str(payload.get("value", ""))
    if row < 1 or col < 1 or row > 1000 or col > 100:
        return JSONResponse({"ok": False, "error": "Invalid cell"}, status_code=400)
    eq = q1("SELECT * FROM trs_equipment WHERE id=?", (equipment_id,))
    if not eq:
        return JSONResponse({"ok": False, "error": "Equipment not found"}, status_code=404)
    if value == "":
        qx("DELETE FROM trs_equipment_sheet_cells WHERE equipment_id=? AND row_index=? AND col_index=?", (equipment_id, row, col))
    else:
        qx("""
            INSERT INTO trs_equipment_sheet_cells (equipment_id,row_index,col_index,value,updated_by,updated_at)
            VALUES (?,?,?,?,?,datetime('now'))
            ON CONFLICT(equipment_id,row_index,col_index) DO UPDATE SET value=excluded.value, updated_by=excluded.updated_by, updated_at=datetime('now')
        """, (equipment_id, row, col, value, user["id"]))
    meta = get_equipment_sheet_matrix(equipment_id, row, col)
    qx("INSERT OR REPLACE INTO trs_equipment_sheets (equipment_id,sheet_name,max_row,max_col,updated_by,updated_at) VALUES (?,COALESCE((SELECT sheet_name FROM trs_equipment_sheets WHERE equipment_id=?),'Live Sheet'),?,?,?,datetime('now'))",
       (equipment_id, equipment_id, meta["max_row"], meta["max_col"], user["id"]))
    source_ok, source_msg = write_equipment_cell_to_source(equipment_id, row, col, value)
    log_activity("sheet_cell_update", user["id"], "trs_equipment", equipment_id, f"Cell {excel_col_name(col)}{row} updated" + (" and saved to original Excel" if source_ok else ""))
    await ws_manager.broadcast({"type":"equipment_sheet_update", "equipment_id": equipment_id, "row": row, "col": col, "value": value, "by": user.get("full_name") or user.get("username")}, "equipment")
    return JSONResponse({"ok": True, "source_saved": source_ok, "source_message": source_msg})


@app.post("/equipment-maintenance/{equipment_id}/sheet/add-row")
async def equipment_sheet_add_row(request: Request, equipment_id: int):
    user, redir = require_user(request)
    if redir: return redir
    sheet = get_equipment_sheet_matrix(equipment_id)
    new_row = sheet["max_row"] + 1
    max_col = sheet["max_col"]
    qx("INSERT OR REPLACE INTO trs_equipment_sheets (equipment_id,sheet_name,max_row,max_col,updated_by,updated_at) VALUES (?,COALESCE((SELECT sheet_name FROM trs_equipment_sheets WHERE equipment_id=?),'Live Sheet'),?,?,?,datetime('now'))",
       (equipment_id, equipment_id, new_row, max_col, user["id"]))
    log_activity("sheet_row_add", user["id"], "trs_equipment", equipment_id, f"Added spreadsheet row {new_row}")
    await ws_manager.broadcast({"type":"equipment_sheet_row_add", "equipment_id": equipment_id}, "equipment")
    return JSONResponse({"ok": True, "row": new_row})


@app.get("/equipment-maintenance/{equipment_id}/sheet/export.xlsx")
async def equipment_sheet_export(request: Request, equipment_id: int):
    user, redir = require_user(request)
    if redir: return redir
    eq = q1("SELECT * FROM trs_equipment WHERE id=?", (equipment_id,))
    if not eq:
        return JSONResponse({"error":"Equipment not found"}, status_code=404)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    wb = Workbook(); ws = wb.active
    sheet = get_equipment_sheet_matrix(equipment_id)
    ws.title = (sheet["meta"].get("sheet_name") if sheet["meta"] else "Maintenance") or "Maintenance"
    for row in sheet["rows"]:
        ws.append(row)
    if ws.max_row >= 1:
        fill = PatternFill("solid", fgColor="0F1F35")
        font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="243A5A")
        for cell in ws[1]:
            cell.fill = fill; cell.font = font; cell.alignment = Alignment(horizontal="center"); cell.border = Border(bottom=thin)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    filename = f"{safe_name(eq.get('category') or 'Equipment')}_{safe_name(eq.get('serial_no') or str(equipment_id))}_live.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition":f"attachment; filename={filename}"})


@app.post("/equipment-maintenance/{equipment_id}/sheet/import-source")
async def equipment_sheet_import_source(request: Request, equipment_id: int):
    user, redir = require_user(request)
    if redir: return redir
    eq = q1("SELECT * FROM trs_equipment WHERE id=?", (equipment_id,))
    source_path = resolve_equipment_source_path(eq)
    if not eq or not source_path or not os.path.exists(source_path):
        return JSONResponse({"ok": False, "error":"No source Excel file"}, status_code=404)
    result = import_equipment_sheet_from_file(equipment_id, source_path, user["id"])
    await ws_manager.broadcast({"type":"equipment_sheet_import", "equipment_id": equipment_id}, "equipment")
    return JSONResponse(result)


@app.get("/equipment-maintenance/{equipment_id}/source")
async def equipment_source_file(request: Request, equipment_id: int):
    user, redir = require_user(request)
    if redir: return redir
    eq = q1("SELECT * FROM trs_equipment WHERE id=?", (equipment_id,))
    source_path = resolve_equipment_source_path(eq)
    if not eq or not source_path or not os.path.exists(source_path):
        return JSONResponse({"error":"Source Excel file not found"}, status_code=404)
    if source_path != (eq.get("source_file_path") or ""):
        qx("UPDATE trs_equipment SET source_file_path=?, updated_at=datetime('now') WHERE id=?", (source_path, equipment_id))
    return FileResponse(source_path, filename=eq.get("source_original_name") or f"equipment_{equipment_id}.xlsx")


@app.get("/equipment-maintenance/{equipment_id}/sheet/download-updated.xlsx")
async def equipment_sheet_download_updated(request: Request, equipment_id: int):
    """Download the original Excel workbook after Live Sync edits have been written into it."""
    user, redir = require_user(request)
    if redir: return redir
    eq = q1("SELECT * FROM trs_equipment WHERE id=?", (equipment_id,))
    source_path = resolve_equipment_source_path(eq)
    if not eq or not source_path or not os.path.exists(source_path):
        return JSONResponse({"error":"Updated source Excel file not found"}, status_code=404)
    return FileResponse(source_path, filename=eq.get("source_original_name") or f"equipment_{equipment_id}_updated.xlsx")


@app.get("/equipment-maintenance/export.xlsx")
async def equipment_export_xlsx(request: Request):
    user, redir = require_user(request)
    if redir: return redir
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return RedirectResponse("/equipment-maintenance/export.csv", status_code=303)
    rows = qa("""
        SELECT e.name, e.category, e.serial_no, e.manufacture_date, e.last_service_date, e.next_service_date, e.status, e.location, e.notes,
               l.service_date, l.service_type, l.performed_by, l.next_due_date, l.notes AS log_notes
        FROM trs_equipment e LEFT JOIN trs_equipment_logs l ON l.equipment_id=e.id
        ORDER BY e.category, e.serial_no, l.service_date DESC
    """)
    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenance Register"
    headers = ["Equipment", "Category", "Serial", "Manufacture Date", "Last Service", "Next Service", "Status", "Location", "Equipment Notes", "Log Date", "Service Type", "Performed By", "Log Next Due", "Log Notes"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(k, "") for k in ["name","category","serial_no","manufacture_date","last_service_date","next_service_date","status","location","notes","service_date","service_type","performed_by","next_due_date","log_notes"]])
    fill = PatternFill("solid", fgColor="0F1F35")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="243A5A")
    for cell in ws[1]:
        cell.fill = fill; cell.font = font; cell.alignment = Alignment(horizontal="center"); cell.border = Border(bottom=thin)
    widths = [26,16,14,18,16,16,16,18,36,16,18,18,16,42]
    for i,w in enumerate(widths, start=1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    bio = io.BytesIO()
    wb.save(bio); bio.seek(0)
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition":"attachment; filename=trs_equipment_maintenance.xlsx"})

@app.get("/equipment-maintenance/export.csv")
async def equipment_export(request: Request):
    user, redir = require_user(request)
    if redir: return redir
    rows = qa("""
        SELECT e.name, e.category, e.serial_no, e.manufacture_date, e.last_service_date, e.next_service_date, e.status, e.location, e.notes,
               l.service_date, l.service_type, l.performed_by, l.next_due_date, l.notes AS log_notes
        FROM trs_equipment e LEFT JOIN trs_equipment_logs l ON l.equipment_id=e.id
        ORDER BY e.name, l.service_date DESC
    """)
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["Equipment", "Category", "Serial", "Manufacture Date", "Last Service", "Next Service", "Status", "Location", "Equipment Notes", "Log Date", "Service Type", "Performed By", "Log Next Due", "Log Notes"])
    for r in rows:
        writer.writerow([r.get(k, "") for k in ["name","category","serial_no","manufacture_date","last_service_date","next_service_date","status","location","notes","service_date","service_type","performed_by","next_due_date","log_notes"]])
    return StreamingResponse(iter([out.getvalue()]), media_type="text/csv", headers={"Content-Disposition":"attachment; filename=trs_equipment_maintenance.csv"})



# ══════════════════════════════════════════════════════════════
#  TRS AI AUTOMATION + NOTIFICATIONS
# ══════════════════════════════════════════════════════════════

def create_notification(level: str, title: str, message: str = "", entity_type: str = "", entity_id: int | None = None, user_id: int | None = None) -> int:
    """Create a platform notification. WebSocket broadcast is done by caller when async context is available."""
    return qx("""
        INSERT INTO trs_notifications (level, title, message, entity_type, entity_id, created_by)
        VALUES (?,?,?,?,?,?)
    """, (level or "info", title or "Notification", message or "", entity_type or "", entity_id, user_id))


def run_trs_ai_automation(user_id: int | None = None) -> dict:
    """Offline-safe automation: no external AI key required.

    What it does now:
    1) Reviews low-confidence job files and suggests/updates detected_path.
    2) Finds overdue / due-soon equipment and creates alerts.
    3) Flags equipment Excel files that are missing from disk.
    4) Creates a clear run summary stored in DB.

    If OpenAI/Firebase is added later, this function is the single place to extend.
    """
    actions = []
    created_notifications = 0

    # 1) Job file organization recommendations
    files = qa("""
        SELECT * FROM trs_job_files
        WHERE COALESCE(detection_confidence,0) < 75 OR COALESCE(detected_path,'')=''
        ORDER BY created_at DESC LIMIT 100
    """)
    for f in files:
        fname = f.get("original_name") or ""
        content = None
        path = f.get("file_path") or ""
        try:
            if path and os.path.exists(path) and file_ext(fname) == "pdf":
                with open(path, "rb") as fh:
                    content = fh.read(2_000_000)
        except Exception:
            content = None
        detected = detect_job_file_path(fname, content)
        det_path = f"{detected['company']}/{detected['job_year']}/{detected['rig']}/{detected['connection_name']}"
        qx("UPDATE trs_job_files SET detection_confidence=?, detected_path=? WHERE id=?", (int(detected.get("confidence") or 0), det_path, f["id"]))
        actions.append({"type":"job_file_ai_detect", "file": fname, "suggested_path": det_path, "confidence": detected.get("confidence")})

    # 2) Equipment maintenance alerts
    overdue = qa("""
        SELECT id, name, category, serial_no, next_service_date
        FROM trs_equipment
        WHERE next_service_date IS NOT NULL AND date(next_service_date) < date('now')
        ORDER BY next_service_date ASC LIMIT 50
    """)
    due = qa("""
        SELECT id, name, category, serial_no, next_service_date
        FROM trs_equipment
        WHERE next_service_date IS NOT NULL AND date(next_service_date) BETWEEN date('now') AND date('now','+30 day')
        ORDER BY next_service_date ASC LIMIT 50
    """)
    for e in overdue:
        title = f"Overdue maintenance: {e.get('name') or e.get('serial_no')}"
        msg = f"Next service date was {e.get('next_service_date')}."
        create_notification("critical", title, msg, "trs_equipment", e["id"], user_id)
        created_notifications += 1
        actions.append({"type":"equipment_overdue", "equipment_id": e["id"], "next_service_date": e.get("next_service_date")})
    for e in due:
        title = f"Maintenance due soon: {e.get('name') or e.get('serial_no')}"
        msg = f"Next service date is {e.get('next_service_date')}."
        create_notification("warning", title, msg, "trs_equipment", e["id"], user_id)
        created_notifications += 1
        actions.append({"type":"equipment_due_soon", "equipment_id": e["id"], "next_service_date": e.get("next_service_date")})

    # 3) Missing Excel source checks
    equipment = qa("SELECT * FROM trs_equipment ORDER BY updated_at DESC LIMIT 200")
    for e in equipment:
        src = resolve_equipment_source_path(e)
        if e.get("source_original_name") and (not src or not os.path.exists(src)):
            title = f"Missing Excel source: {e.get('name') or e.get('serial_no')}"
            msg = f"Original file {e.get('source_original_name')} could not be found. Re-import the equipment file."
            create_notification("warning", title, msg, "trs_equipment", e["id"], user_id)
            created_notifications += 1
            actions.append({"type":"missing_excel_source", "equipment_id": e["id"], "file": e.get("source_original_name")})

    summary = f"AI automation completed: {len(actions)} actions, {created_notifications} notifications."
    run_id = qx("""
        INSERT INTO trs_ai_runs (run_type, status, summary, details_json, created_by)
        VALUES (?,?,?,?,?)
    """, ("manual", "completed", summary, json.dumps(actions, ensure_ascii=False), user_id))
    if user_id:
        log_activity("ai_automation_run", user_id, "trs_ai_run", run_id, summary)
    return {"ok": True, "run_id": run_id, "summary": summary, "actions": actions, "notifications": created_notifications}


@app.websocket("/ws/trs/{channel}")
async def trs_ws(ws: WebSocket, channel: str):
    # Best-effort user identification from session cookie for presence.
    user = None
    try:
        token = ws.cookies.get("trs_session") or ws.cookies.get("jam_session")
        sess = session_store.get(token) if token else None
        user = sess.get("user") if sess else None
    except Exception:
        user = None
    await ws_manager.connect(ws, channel, user)
    try:
        while True:
            # keep connection open; client may send ping/presence payloads
            await ws.receive_text()
    except WebSocketDisconnect:
        ws_manager.disconnect(ws, channel)
        if channel not in ('job_files','equipment','scada'):
            await ws_manager.broadcast({"type":"presence", "event":"leave", "channel":channel, "clients":ws_manager.client_count()}, channel)



@app.get("/healthz")
async def healthz():
    """Deployment health check for Render/Railway/Fly/VPS."""
    try:
        db_ok = q1("SELECT 1 ok")
        return {"ok": True, "app": APP_NAME, "version": APP_VERSION, "db": bool(db_ok), "ws_clients": ws_manager.client_count()}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@app.get("/api/realtime/status")
async def realtime_status(request: Request):
    user, redir = require_user(request)
    if redir: return redir
    return JSONResponse(ws_manager.status())


@app.get("/api/notifications")
async def api_notifications(request: Request, unread_only: int = 0, limit: int = 30):
    user, redir = require_user(request)
    if redir: return redir
    where = "WHERE is_read=0" if unread_only else ""
    rows = qa(f"SELECT * FROM trs_notifications {where} ORDER BY created_at DESC LIMIT ?", (min(max(limit, 1), 100),))
    return JSONResponse({"ok": True, "notifications": rows})


@app.post("/api/notifications/{notification_id}/read")
async def api_notification_read(request: Request, notification_id: int):
    user, redir = require_user(request)
    if redir: return redir
    qx("UPDATE trs_notifications SET is_read=1 WHERE id=?", (notification_id,))
    return JSONResponse({"ok": True})



@app.post("/api/notifications/read-all")
async def api_notifications_read_all(request: Request):
    user, redir = require_user(request)
    if redir: return redir
    qx("UPDATE trs_notifications SET is_read=1 WHERE is_read=0", ())
    return JSONResponse({"ok": True})

@app.get("/ai-automation", response_class=HTMLResponse)
async def ai_automation_page(request: Request):
    user, redir = require_user(request)
    if redir: return redir
    runs = qa("SELECT * FROM trs_ai_runs ORDER BY created_at DESC LIMIT 20")
    notes = qa("SELECT * FROM trs_notifications ORDER BY created_at DESC LIMIT 40")
    return render(request, "ai_automation.html", {"active":"ai_automation", "runs":runs, "notifications":notes})


@app.post("/api/ai-automation/run")
async def api_ai_automation_run(request: Request):
    user, redir = require_user(request)
    if redir: return redir
    result = run_trs_ai_automation(user.get("id"))
    await ws_manager.broadcast({"type":"ai_automation_run", "summary": result["summary"], "notifications": result["notifications"]}, "all")
    return JSONResponse(result)


@app.get("/activity", response_class=HTMLResponse)
async def activity_page(request: Request, entity_type: str = "", q: str = "", limit: int = 100):
    user, redir = require_user(request)
    if redir: return redir
    limit = min(max(limit, 10), 500)
    where, params = [], []
    if entity_type:
        where.append("a.entity_type=?"); params.append(entity_type)
    if q:
        where.append("(a.action LIKE ? OR a.detail LIKE ? OR u.username LIKE ? OR u.full_name LIKE ?)"); params.extend([f"%{q}%"]*4)
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    rows = qa(f"""
        SELECT a.*, u.username, u.full_name, u.role
        FROM activity_log a LEFT JOIN users u ON u.id=a.user_id
        {where_sql}
        ORDER BY a.created_at DESC LIMIT {limit}
    """, tuple(params))
    activity_counts = {
        "upload_create": sum(1 for r in rows if ("upload" in (r.get("action") or "") or "create" in (r.get("action") or ""))),
        "update_edit": sum(1 for r in rows if ("update" in (r.get("action") or "") or "edit" in (r.get("action") or ""))),
        "login": sum(1 for r in rows if (r.get("action") or "") == "login"),
    }
    return render(request, "activity.html", {
        "active":"activity", "rows": rows,
        "activity_counts": activity_counts,
        "filters": {"entity_type": entity_type, "q": q, "limit": limit}
    })


# ══════════════════════════════════════════════════════════════
#  JOBS
# ══════════════════════════════════════════════════════════════

@app.get("/jobs", response_class=HTMLResponse)
async def jobs_list(request: Request, status: str = "", q: str = ""):
    user, redir = require_user(request)
    if redir: return redir

    where = "WHERE 1=1"
    params = []
    if status:
        where += " AND j.status=?"; params.append(status)
    if q:
        where += " AND (j.job_number LIKE ? OR j.customer LIKE ? OR j.rig LIKE ?)"
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]

    jobs = qa(f"""
        SELECT j.id, j.job_number, j.customer, j.rig, j.well, j.status, j.start_date,
               d.code device_code, d.name device_name,
               m.total_joints, m.ft_mean, m.rerun_rate, m.computed_at,
               (SELECT COUNT(*) FROM mtt_joint_data WHERE job_id=j.id) joint_count
        FROM jobs j
        LEFT JOIN devices d ON d.id=j.assigned_device_id
        LEFT JOIN mtt_job_summary m ON m.job_id=j.id
        {where}
        ORDER BY j.created_at DESC
    """, tuple(params))

    devices = qa("SELECT id, code, name FROM devices WHERE status='Available' OR status='In Job' ORDER BY code")
    return render(request, "jobs.html", {
        "active": "jobs", "jobs": jobs, "devices": devices,
        "filter_status": status, "filter_q": q,
        "job_statuses": JOB_STATUSES,
    })


@app.post("/jobs/create")
async def job_create(request: Request,
                     job_number: str = Form(...),
                     customer: str = Form(""),
                     rig: str = Form(""),
                     well: str = Form(""),
                     field: str = Form(""),
                     country: str = Form(""),
                     device_id: str = Form(""),
                     status: str = Form("Planned"),
                     start_date: str = Form(""),
                     notes: str = Form("")):
    user, redir = require_user(request)
    if redir: return redir
    if not rbac.can(user, "jobs.create"):
        return RedirectResponse("/?error=no_permission", status_code=303)

    # Check duplicate job number
    existing = q1("SELECT id FROM jobs WHERE job_number=?", (job_number,))
    if existing:
        devices = qa("SELECT id, code, name FROM devices ORDER BY code")
        jobs = qa("SELECT j.*, d.code device_code FROM jobs j LEFT JOIN devices d ON d.id=j.assigned_device_id ORDER BY j.created_at DESC")
        return render(request, "jobs.html", {
            "active": "jobs", "jobs": jobs, "devices": devices,
            "error": f"Job number '{job_number}' already exists",
            "job_statuses": JOB_STATUSES,
        })

    did = int(device_id) if device_id and device_id.isdigit() else None
    job_id = qx("""
        INSERT INTO jobs (job_number, customer, rig, well, field, country,
                          assigned_device_id, status, start_date, notes, created_by)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (job_number, customer, rig, well, field, country, did, status,
          start_date or None, notes, user["id"]))

    if did:
        qx("UPDATE devices SET status='In Job' WHERE id=?", (did,))

    log_activity("job_create", user["id"], "job", job_id, f"Created job {job_number}")
    try:
        jr = q1("SELECT * FROM jobs WHERE id=?", (job_id,))
        if jr: sync_job_to_fs(dict(jr))
    except Exception as e: logger.debug(f"Ignored non-critical error: {e}")
    return RedirectResponse(f"/jobs/{job_id}", status_code=303)


@app.get("/jobs/{job_id}", response_class=HTMLResponse)
async def job_detail(job_id: int, request: Request):
    user, redir = require_user(request)
    if redir: return redir

    job = q1("""SELECT j.*, d.code device_code, d.name device_name
                FROM jobs j LEFT JOIN devices d ON d.id=j.assigned_device_id
                WHERE j.id=?""", (job_id,))
    if not job:
        return RedirectResponse("/jobs", status_code=303)

    summary = q1("""SELECT id,job_id,file_id,pipe_type,tong_model,total_joints,ok_count,
                           rerun_count,rerun_rate,outlier_count,ft_mean,ft_std,ft_min,ft_max,
                           turns_mean,rpm_mean,dt_mean,low_rpm_count,outliers_high,outliers_low,
                           reruns,fast_joints,slow_joints,stats_json,computed_at
                    FROM mtt_job_summary WHERE job_id=?""", (job_id,))
    joints = qa("""SELECT * FROM mtt_joint_data WHERE job_id=?
                   ORDER BY joint_num ASC, run_number ASC""", (job_id,))

    if summary:
        for f in ["outliers_high","outliers_low","reruns","fast_joints","slow_joints","stats_json"]:
            if summary.get(f) and isinstance(summary[f], str):
                try: summary[f] = json.loads(summary[f])
                except Exception as e: logger.debug(f"Ignored non-critical error: {e}")

    files = qa("SELECT * FROM job_data_files WHERE job_id=? ORDER BY created_at DESC", (job_id,))
    ai = q1("SELECT * FROM job_ai_insights WHERE job_id=?", (job_id,))

    # job rating
    job_rating = None
    if summary and joints:
        try:
            from services.job_analysis_pdf import compute_job_rating
            sj = summary.get("stats_json") or {}
            specs = {k: (sj if isinstance(sj,dict) else {}).get(k) for k in
                     ["max_torque","opt_torque","min_torque","high_shoulder","low_shoulder"]}
            job_rating = compute_job_rating([dict(j) for j in joints], dict(summary), specs)
        except Exception:
            pass

    devices = qa("SELECT id, code, name, status FROM devices ORDER BY code")
    return render(request, "job_detail.html", {
        "active": "jobs", "job": job, "summary": summary, "joints": joints,
        "files": files, "ai": ai, "job_rating": job_rating, "devices": devices,
        "job_statuses": JOB_STATUSES,
    })


@app.post("/jobs/{job_id}/update-status")
async def job_update_status(job_id: int, request: Request, status: str = Form(...)):
    user, redir = require_user(request)
    if redir: return redir
    qx("UPDATE jobs SET status=? WHERE id=?", (status, job_id))
    log_activity("job_status", user["id"], "job", job_id, f"Status → {status}")
    return RedirectResponse(f"/jobs/{job_id}", status_code=303)


@app.post("/jobs/{job_id}/assign-device")
async def job_assign_device(job_id: int, request: Request, device_id: str = Form("")):
    user, redir = require_user(request)
    if redir: return redir
    did = int(device_id) if device_id and device_id.isdigit() else None
    qx("UPDATE jobs SET assigned_device_id=? WHERE id=?", (did, job_id))
    if did:
        qx("UPDATE devices SET status='In Job' WHERE id=?", (did,))
    log_activity("job_assign_device", user["id"], "job", job_id, f"Device → {did}")
    return RedirectResponse(f"/jobs/{job_id}", status_code=303)



@app.post("/jobs/{job_id}/delete")
async def job_delete(job_id: int, request: Request):
    user, redir = require_role(request, ["Admin", "Manager"])
    if redir: return redir
    job = q1("SELECT * FROM jobs WHERE id=?", (job_id,))
    if job:
        qx("DELETE FROM jobs WHERE id=?", (job_id,))
        log_activity("job_delete", user["id"], "job", job_id, f"Deleted job {job.get('job_number')}")
        await ws_manager.broadcast({"type":"job_delete", "job_id": job_id}, "all")
    return RedirectResponse("/jobs", status_code=303)

# ══════════════════════════════════════════════════════════════
#  MTT UPLOAD & JOB ANALYSIS
# ══════════════════════════════════════════════════════════════

@app.post("/api/upload-mtt/{job_id}")
async def upload_mtt(job_id: int, request: Request, file: UploadFile = File(...)):
    user = get_user(request)
    if not user:
        return JSONResponse({"ok": False, "error": "Not authorized"}, 401)

    content = await file.read()
    if len(content) > MAX_UPLOAD_MB * 1024 * 1024:
        return JSONResponse({"ok": False, "error": f"File exceeds {MAX_UPLOAD_MB}MB limit"}, 400)

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("pdf", "csv"):
        return JSONResponse({"ok": False, "error": "Only PDF or CSV files accepted"}, 400)

    try:
        if ext == "csv":
            from services.mtt_csv_parser import parse_mtt_csv
            mtt = parse_mtt_csv(content)
        else:
            from services.mtt_parser import parse_mtt_pdf
            mtt = parse_mtt_pdf(content)

        if not mtt or not mtt.get("is_mtt"):
            return JSONResponse({"ok": False, "error": "File is not a valid MTT report. Please upload a PDF or CSV exported from the MTT system."}, 400)

        joints = mtt["joints"]
        stats  = mtt["stats"]
        header = mtt["header"]

        # Save to disk
        safe_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:6]}.{ext}"
        try:
            folder = UPLOADS_DIR / "mtt"
            folder.mkdir(parents=True, exist_ok=True)
            (folder / safe_name).write_bytes(content)
        except Exception as e:
            logger.warning(f"Disk save failed (using DB blob): {e}")

        # Register file
        file_id = qx("""
            INSERT INTO job_data_files
                (job_id, file_name, file_path, file_type, row_count, parsed, uploaded_by)
            VALUES (?,?,?,?,?,1,?)
        """, (job_id, file.filename, f"/uploads/mtt/{safe_name}", ext, len(joints), user["id"]))

        # Store joint data
        conn = get_conn()
        conn.execute("DELETE FROM mtt_joint_data WHERE job_id=?", (job_id,))
        for j in joints:
            conn.execute("""
                INSERT INTO mtt_joint_data
                    (job_id,file_id,joint_label,joint_num,run_number,is_rerun,lot,datetime,
                     final_torque,final_turns,shoulder_torque,shoulder_turns,shoulder_rpm,
                     delta_torque,delta_turns,comment)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (job_id, file_id, j["joint_label"], j["joint_num"], j["run_number"],
                  int(j["is_rerun"]), j.get("lot",""), j["datetime"],
                  j["final_torque"], j["final_turns"], j["shoulder_torque"],
                  j["shoulder_turns"], j["shoulder_rpm"], j["delta_torque"],
                  j["delta_turns"], j.get("comment","")))

        # Store summary
        conn.execute("""
            INSERT INTO mtt_job_summary
                (job_id,file_id,pipe_type,tong_model,total_joints,ok_count,rerun_count,
                 rerun_rate,outlier_count,ft_mean,ft_std,ft_min,ft_max,turns_mean,rpm_mean,
                 dt_mean,low_rpm_count,outliers_high,outliers_low,reruns,fast_joints,
                 slow_joints,stats_json,pdf_blob,computed_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
            ON CONFLICT(job_id) DO UPDATE SET
                file_id=excluded.file_id, pipe_type=excluded.pipe_type,
                tong_model=excluded.tong_model, total_joints=excluded.total_joints,
                ok_count=excluded.ok_count, rerun_count=excluded.rerun_count,
                rerun_rate=excluded.rerun_rate, ft_mean=excluded.ft_mean,
                ft_std=excluded.ft_std, ft_min=excluded.ft_min, ft_max=excluded.ft_max,
                turns_mean=excluded.turns_mean, rpm_mean=excluded.rpm_mean,
                outliers_high=excluded.outliers_high, outliers_low=excluded.outliers_low,
                reruns=excluded.reruns, fast_joints=excluded.fast_joints,
                slow_joints=excluded.slow_joints, stats_json=excluded.stats_json,
                pdf_blob=excluded.pdf_blob, computed_at=datetime('now')
        """, (
            job_id, file_id, header.get("pipe_type",""), header.get("tong_model",""),
            stats.get("total_joints",0), stats.get("ok_count",0),
            stats.get("rerun_count",0), stats.get("rerun_rate",0.0),
            stats.get("outlier_count",0),
            stats.get("final_torque",{}).get("mean"),
            stats.get("final_torque",{}).get("std"),
            stats.get("final_torque",{}).get("min"),
            stats.get("final_torque",{}).get("max"),
            stats.get("final_turns",{}).get("mean"),
            stats.get("shoulder_rpm",{}).get("mean"),
            stats.get("delta_torque",{}).get("mean"),
            len(stats.get("low_rpm_joints",[])),
            json.dumps(stats.get("outliers_high",[])),
            json.dumps(stats.get("outliers_low",[])),
            json.dumps(stats.get("reruns",[])),
            json.dumps(stats.get("fast_joints",[])),
            json.dumps(stats.get("slow_joints",[])),
            json.dumps({**stats, **{
                "max_torque":   header.get("max_torque"),
                "opt_torque":   header.get("opt_torque"),
                "min_torque":   header.get("min_torque"),
                "high_shoulder":header.get("high_shoulder"),
                "low_shoulder": header.get("low_shoulder"),
            }}),
            content if ext == "pdf" else None,
        ))
        conn.commit()
        conn.close()

        logger.info(f"MTT upload OK: job={job_id} joints={len(joints)}")
        log_activity("mtt_upload", user["id"], "job", job_id,
                     f"MTT {ext.upper()} uploaded: {len(joints)} joints")

        # Sync summary to Firestore
        try:
            sr = q1("SELECT job_id,pipe_type,tong_model,total_joints,ok_count,rerun_count,rerun_rate,outlier_count,ft_mean,ft_std,ft_min,ft_max,turns_mean,rpm_mean,computed_at FROM mtt_job_summary WHERE job_id=?", (job_id,))
            if sr: sync_summary_to_fs(dict(sr))
        except Exception as ex:
            logger.debug(f"FS summary sync: {ex}")

        # Upload original PDF to Cloudinary
        if ext == "pdf" and cl_enabled():
            try:
                cl_url = upload_mtt_original(content, job_id, file.filename)
                if cl_url:
                    qx("UPDATE job_data_files SET file_path=? WHERE id=?", (cl_url, file_id))
                    logger.info(f"MTT PDF on Cloudinary: {cl_url}")
            except Exception as ex:
                logger.debug(f"Cloudinary MTT: {ex}")

        return JSONResponse({
            "ok": True,
            "joints": len(joints),
            "reruns": stats.get("rerun_count", 0),
            "pipe_type": header.get("pipe_type",""),
            "ft_mean": stats.get("final_torque",{}).get("mean"),
            "max_torque": header.get("max_torque"),
            "opt_torque": header.get("opt_torque"),
            "min_torque": header.get("min_torque"),
            "redirect": f"/jobs/{job_id}",
        })

    except Exception as e:
        logger.error(f"MTT upload error: {e}", exc_info=True)
        return JSONResponse({"ok": False, "error": f"Analysis failed: {str(e)}"}, 500)


@app.post("/api/repair-blob/{job_id}")
async def repair_blob(job_id: int, request: Request, file: UploadFile = File(...)):
    """Re-upload original MTT PDF to restore graph extraction capability."""
    user = get_user(request)
    if not user:
        return JSONResponse({"ok": False, "error": "Not authorized"}, 401)
    content = await file.read()
    if not content or not file.filename.lower().endswith(".pdf"):
        return JSONResponse({"ok": False, "error": "Please upload the original MTT PDF"}, 400)
    row = q1("SELECT job_id FROM mtt_job_summary WHERE job_id=?", (job_id,))
    if not row:
        return JSONResponse({"ok": False, "error": "No MTT data found — upload MTT file first"}, 400)
    conn = get_conn()
    conn.execute("UPDATE mtt_job_summary SET pdf_blob=? WHERE job_id=?", (content, job_id))
    conn.commit(); conn.close()
    return JSONResponse({"ok": True, "size": len(content), "message": "PDF blob saved — graphs restored"})


# ══════════════════════════════════════════════════════════════
#  JOB ANALYSIS PAGE (Full MTT Analysis)
# ══════════════════════════════════════════════════════════════

@app.get("/analysis", response_class=HTMLResponse)
async def analysis_page(request: Request, job_id: int = None):
    user, redir = require_user(request)
    if redir: return redir
    if not rbac.can(user, "analysis.view"):
        from fastapi.responses import HTMLResponse as HR
        return HR("""<meta http-equiv="refresh" content="2;url=/">
            <div style="font-family:sans-serif;background:#050d1a;color:#e2e8f0;padding:40px;text-align:center">
            🔒 Access denied — analysis requires Engineer role or above</div>""", 403)

    all_jobs = qa("""
        SELECT j.id, j.job_number, j.customer, j.rig, j.status,
               m.total_joints, m.ft_mean, m.rerun_rate, m.computed_at
        FROM jobs j
        LEFT JOIN mtt_job_summary m ON m.job_id=j.id
        ORDER BY j.created_at DESC
    """)

    selected_job = summary = job_rating = None
    joints = []

    if job_id:
        selected_job = q1("""SELECT j.*, d.code device_code, d.name device_name
                              FROM jobs j LEFT JOIN devices d ON d.id=j.assigned_device_id
                              WHERE j.id=?""", (job_id,))
        summary = q1("""SELECT id,job_id,file_id,pipe_type,tong_model,total_joints,ok_count,
                               rerun_count,rerun_rate,outlier_count,ft_mean,ft_std,ft_min,ft_max,
                               turns_mean,rpm_mean,dt_mean,low_rpm_count,outliers_high,outliers_low,
                               reruns,fast_joints,slow_joints,stats_json,computed_at
                        FROM mtt_job_summary WHERE job_id=?""", (job_id,))
        joints = qa("""SELECT * FROM mtt_joint_data WHERE job_id=?
                       ORDER BY joint_num ASC, run_number ASC""", (job_id,))
        if summary:
            for f in ["outliers_high","outliers_low","reruns","fast_joints","slow_joints","stats_json"]:
                if summary.get(f) and isinstance(summary[f], str):
                    try: summary[f] = json.loads(summary[f])
                    except Exception as e: logger.debug(f"Ignored non-critical error: {e}")
        if summary and joints:
            try:
                from services.job_analysis_pdf import compute_job_rating
                sj = summary.get("stats_json") or {}
                specs = {k: (sj if isinstance(sj,dict) else {}).get(k) for k in
                         ["max_torque","opt_torque","min_torque","high_shoulder","low_shoulder"]}
                job_rating = compute_job_rating([dict(j) for j in joints], dict(summary), specs)
            except Exception:
                pass

    return render(request, "analysis.html", {
        "active": "analysis",
        "all_jobs": all_jobs,
        "selected_job": selected_job,
        "selected_job_id": job_id,
        "summary": summary,
        "joints": joints,
        "job_rating": job_rating,
    })


# ══════════════════════════════════════════════════════════════
#  FLEET ANALYSIS
# ══════════════════════════════════════════════════════════════

@app.get("/fleet", response_class=HTMLResponse)
async def fleet_page(request: Request):
    user, redir = require_user(request)
    if redir: return redir

    jobs = qa("""
        SELECT j.id, j.job_number, j.customer, j.rig, j.status, j.start_date,
               d.code device_code, d.name device_name,
               m.pipe_type, m.tong_model, m.total_joints, m.ok_count,
               m.rerun_count, m.rerun_rate, m.ft_mean, m.ft_std, m.ft_min, m.ft_max,
               m.outlier_count, m.low_rpm_count, m.computed_at
        FROM mtt_job_summary m
        JOIN jobs j ON j.id=m.job_id
        LEFT JOIN devices d ON d.id=j.assigned_device_id
        ORDER BY m.computed_at DESC
    """)

    # Fleet-level stats
    if jobs:
        ft_means = [j["ft_mean"] for j in jobs if j["ft_mean"]]
        fleet_stats = {
            "total_jobs": len(jobs),
            "total_joints": sum(j["total_joints"] or 0 for j in jobs),
            "avg_rerun_rate": round(sum(j["rerun_rate"] or 0 for j in jobs) / len(jobs), 1),
            "avg_ft_mean": round(sum(ft_means) / len(ft_means), 1) if ft_means else 0,
        }
    else:
        fleet_stats = {"total_jobs": 0, "total_joints": 0, "avg_rerun_rate": 0, "avg_ft_mean": 0}

    return render(request, "fleet.html", {
        "active": "fleet", "jobs": jobs, "fleet_stats": fleet_stats,
    })


# ══════════════════════════════════════════════════════════════
#  PDF REPORT DOWNLOAD
# ══════════════════════════════════════════════════════════════

@app.get("/api/jobs/{job_id}/report")
async def download_report(job_id: int, request: Request, mode: str = "flagged"):
    user = get_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)

    job = q1("""SELECT j.*, d.name device_name, d.code device_code
                FROM jobs j LEFT JOIN devices d ON d.id=j.assigned_device_id
                WHERE j.id=?""", (job_id,))
    if not job:
        return Response("Job not found", status_code=404)

    summary = q1("""SELECT id,job_id,file_id,pipe_type,tong_model,total_joints,ok_count,
                           rerun_count,rerun_rate,outlier_count,ft_mean,ft_std,ft_min,ft_max,
                           turns_mean,rpm_mean,dt_mean,low_rpm_count,outliers_high,outliers_low,
                           reruns,fast_joints,slow_joints,stats_json,computed_at
                    FROM mtt_job_summary WHERE job_id=?""", (job_id,))
    if not summary:
        return Response("No MTT data for this job. Please upload an MTT report first.", status_code=404)

    joints = qa("SELECT * FROM mtt_joint_data WHERE job_id=? ORDER BY joint_num ASC, run_number ASC", (job_id,))
    if not joints:
        return Response("No joint data found.", status_code=404)

    if summary.get("stats_json") and isinstance(summary["stats_json"], str):
        try: summary["stats_json"] = json.loads(summary["stats_json"])
        except Exception as e: logger.debug(f"Ignored non-critical error: {e}")

    mode = mode if mode in ("full","flagged","critical","issues") else "flagged"

    try:
        orig_pdf = None
        page_map = None
        try:
            blob_row = q1("SELECT pdf_blob FROM mtt_job_summary WHERE job_id=?", (job_id,))
            if blob_row and blob_row.get("pdf_blob"):
                orig_pdf = bytes(blob_row["pdf_blob"])
                from services.graph_extractor import build_joint_page_map
                page_map = build_joint_page_map(orig_pdf)
        except Exception:
            pass

        from services.job_analysis_pdf import build_job_analysis_pdf
        pdf_bytes = build_job_analysis_pdf(
            job=dict(job), summary=dict(summary),
            joints=[dict(j) for j in joints],
            mode=mode, pdf_bytes=orig_pdf, page_map=page_map,
        )
    except Exception as e:
        logger.error(f"Report generation error: {e}", exc_info=True)
        return Response(f"PDF generation failed: {str(e)}", status_code=500)

    fn = f"TRS_Report_{job['job_number'].replace('/','_')}_{mode}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})


# ══════════════════════════════════════════════════════════════
#  DEVICES
# ══════════════════════════════════════════════════════════════

@app.get("/devices", response_class=HTMLResponse)
async def devices_list(request: Request):
    user, redir = require_user(request)
    if redir: return redir

    devices = qa("""
        SELECT d.*,
               (SELECT COUNT(*) FROM jobs WHERE assigned_device_id=d.id) total_jobs,
               (SELECT COUNT(*) FROM jobs WHERE assigned_device_id=d.id AND status='Active') active_jobs
        FROM devices d ORDER BY d.code
    """)
    return render(request, "devices.html", {
        "active": "devices", "devices": devices,
        "device_categories": DEVICE_CATEGORIES,
        "device_statuses": DEVICE_STATUSES,
    })


@app.post("/devices/create")
async def device_create(request: Request,
                        code: str = Form(...),
                        name: str = Form(...),
                        category: str = Form("Torque Turn System"),
                        model: str = Form(""),
                        serial_no: str = Form(""),
                        location: str = Form(""),
                        notes: str = Form("")):
    user, redir = require_user(request)
    if redir: return redir

    existing = q1("SELECT id FROM devices WHERE code=?", (code,))
    if existing:
        return RedirectResponse("/devices?error=duplicate_code", status_code=303)

    did = qx("""INSERT INTO devices (code, name, category, model, serial_no, location, notes)
               VALUES (?,?,?,?,?,?,?)""",
             (code, name, category, model, serial_no, location, notes))
    log_activity("device_create", user["id"], "device", did, f"Created device {code}")
    try:
        dr = q1("SELECT * FROM devices WHERE id=?", (did,))
        if dr: sync_device_to_fs(dict(dr))
    except Exception as e: logger.debug(f"Ignored non-critical error: {e}")
    return RedirectResponse("/devices", status_code=303)


@app.get("/devices/{device_id}", response_class=HTMLResponse)
async def device_detail(device_id: int, request: Request):
    user, redir = require_user(request)
    if redir: return redir

    device = q1("SELECT * FROM devices WHERE id=?", (device_id,))
    if not device:
        return RedirectResponse("/devices", status_code=303)

    jobs = qa("""
        SELECT j.id, j.job_number, j.customer, j.rig, j.status, j.start_date,
               m.total_joints, m.ft_mean, m.rerun_rate
        FROM jobs j LEFT JOIN mtt_job_summary m ON m.job_id=j.id
        WHERE j.assigned_device_id=?
        ORDER BY j.created_at DESC LIMIT 20
    """, (device_id,))

    # Device torque history across jobs
    torque_history = qa("""
        SELECT j.job_number, m.ft_mean, m.ft_min, m.ft_max, m.rerun_rate, m.total_joints
        FROM mtt_job_summary m JOIN jobs j ON j.id=m.job_id
        WHERE j.assigned_device_id=?
        ORDER BY m.computed_at DESC LIMIT 15
    """, (device_id,))

    return render(request, "device_detail.html", {
        "active": "devices", "device": device, "jobs": jobs,
        "torque_history": torque_history,
        "device_statuses": DEVICE_STATUSES,
    })


@app.post("/devices/{device_id}/update-status")
async def device_update_status(device_id: int, request: Request, status: str = Form(...)):
    user, redir = require_user(request)
    if redir: return redir
    qx("UPDATE devices SET status=? WHERE id=?", (status, device_id))
    log_activity("device_status", user["id"], "device", device_id, f"Status → {status}")
    return RedirectResponse(f"/devices/{device_id}", status_code=303)


@app.post("/devices/{device_id}/edit")
async def device_edit(device_id: int, request: Request,
                      code: str = Form(...), name: str = Form(...),
                      category: str = Form("Torque Turn System"),
                      model: str = Form(""), serial_no: str = Form(""),
                      location: str = Form(""), notes: str = Form("")):
    user, redir = require_user(request)
    if redir: return redir
    if not rbac.can(user, "devices.edit"):
        return RedirectResponse(f"/devices/{device_id}", status_code=303)
    qx("""UPDATE devices SET code=?,name=?,category=?,model=?,
             serial_no=?,location=?,notes=? WHERE id=?""",
       (code.strip().upper(), name.strip(), category, model, serial_no, location, notes, device_id))
    log_activity("device_edit", user["id"], "device", device_id, f"Edited device {code}")
    return RedirectResponse(f"/devices/{device_id}", status_code=303)



@app.post("/devices/{device_id}/delete")
async def device_delete(device_id: int, request: Request):
    user, redir = require_role(request, ["Admin", "Manager"])
    if redir: return redir
    device = q1("SELECT * FROM devices WHERE id=?", (device_id,))
    if device:
        qx("DELETE FROM devices WHERE id=?", (device_id,))
        log_activity("device_delete", user["id"], "device", device_id, f"Deleted device {device.get('code')}")
        await ws_manager.broadcast({"type":"device_delete", "device_id": device_id}, "all")
    return RedirectResponse("/devices", status_code=303)

# ══════════════════════════════════════════════════════════════
#  SCADA LIVE
# ══════════════════════════════════════════════════════════════

@app.get("/scada", response_class=HTMLResponse)
async def scada_page(request: Request):
    user, redir = require_user(request)
    if redir: return redir

    devices = qa("SELECT id, code, name, status FROM devices ORDER BY code")
    active_jobs = qa("""
        SELECT j.id, j.job_number, j.customer, j.rig,
               d.id device_id, d.code device_code, d.name device_name
        FROM jobs j JOIN devices d ON d.id=j.assigned_device_id
        WHERE j.status='Active'
        ORDER BY j.start_date DESC
    """)

    # Last readings per device
    last_readings = qa("""
        SELECT r.*, d.code device_code
        FROM scada_readings r JOIN devices d ON d.id=r.device_id
        ORDER BY r.ts DESC LIMIT 20
    """)

    return render(request, "scada.html", {
        "active": "scada", "devices": devices,
        "active_jobs": active_jobs, "last_readings": last_readings,
    })


@app.post("/api/scada/push")
async def scada_push(request: Request):
    """Push live reading from device (called by device firmware or MQTT bridge)."""
    body = await request.json()
    device_id = body.get("device_id")
    job_id    = body.get("job_id")
    ts        = body.get("ts") or datetime.now().isoformat()

    rid = qx("""
        INSERT INTO scada_readings (device_id, job_id, ts, torque, turns, rpm, pressure, temperature, weight, source)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (device_id, job_id, ts,
          body.get("torque"), body.get("turns"), body.get("rpm"),
          body.get("pressure"), body.get("temperature"), body.get("weight"),
          body.get("source","live")))

    # Sync to Firestore
    try:
        sync_scada_to_fs({
            "device_id": device_id, "job_id": job_id, "ts": ts,
            "torque": torque, "turns": turns, "rpm": rpm,
            "pressure": pressure, "temperature": temperature, "weight": weight,
        }, device_id or 0)
    except Exception as e: logger.debug(f"Ignored non-critical error: {e}")

    return JSONResponse({"ok": True, "id": rid})


@app.get("/api/scada/latest")
async def scada_latest(request: Request, device_id: int = None, limit: int = 100):
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Not authorized"}, 401)

    where = "WHERE 1=1"
    params = []
    if device_id:
        where += " AND r.device_id=?"; params.append(device_id)

    readings = qa(f"""
        SELECT r.*, d.code device_code, d.name device_name
        FROM scada_readings r LEFT JOIN devices d ON d.id=r.device_id
        {where}
        ORDER BY r.ts DESC LIMIT ?
    """, tuple(params) + (limit,))

    return JSONResponse([dict(r) for r in readings])




# ══════════════════════════════════════════════════════════════
#  WEBSOCKET — SCADA Real-time stream
# ══════════════════════════════════════════════════════════════

@app.websocket("/ws/scada/{channel}")
async def scada_ws(websocket: WebSocket, channel: str):
    """WebSocket endpoint — channel = device_id or 'all'."""
    await ws_manager.connect(websocket, channel)
    try:
        # Send last 30 readings immediately on connect
        if channel != "all" and channel.isdigit():
            rows = qa(
                "SELECT r.*,d.code device_code FROM scada_readings r "
                "LEFT JOIN devices d ON d.id=r.device_id "
                "WHERE r.device_id=? ORDER BY r.ts DESC LIMIT 30",
                (int(channel),)
            )
        else:
            rows = qa(
                "SELECT r.*,d.code device_code FROM scada_readings r "
                "LEFT JOIN devices d ON d.id=r.device_id "
                "ORDER BY r.ts DESC LIMIT 30"
            )
        await websocket.send_text(json.dumps({
            "type": "history",
            "readings": [dict(r) for r in reversed(rows)]
        }, default=str))

        # Keep connection alive
        while True:
            try:
                await websocket.receive_text()
            except Exception:
                break
    except WebSocketDisconnect:
        pass
    finally:
        ws_manager.disconnect(websocket, channel)


# ══════════════════════════════════════════════════════════════
#  ESP32 PUSH ENDPOINT — No auth (API key based)
# ══════════════════════════════════════════════════════════════

# Simple API key check for ESP devices — set in settings or env
ESP_API_KEY = os.getenv("TRS_ESP_KEY", "trs-esp-key-2026")

@app.post("/api/esp/push")
async def esp_push(request: Request):
    """
    ESP32 pushes readings here.
    Headers: X-ESP-Key: <key>
    Body JSON:
    {
      "device_code": "TRS-001",   // or device_id int
      "torque": 1250.5,
      "turns": 3.142,
      "rpm": 4.2,
      "pressure": 0,
      "temperature": 0,
      "weight": 0,
      "joint_num": 42,             // optional
      "job_id": 1                  // optional
    }
    """
    # Validate API key
    key = request.headers.get("X-ESP-Key", "")
    if key != ESP_API_KEY:
        return JSONResponse({"ok": False, "error": "Invalid API key"}, 401)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "Invalid JSON"}, 400)

    # Resolve device
    device_id = body.get("device_id")
    device_code = body.get("device_code", "")
    device_name = ""

    if not device_id and device_code:
        dev = q1("SELECT id, name FROM devices WHERE code=?", (device_code,))
        if dev:
            device_id = dev["id"]
            device_name = dev["name"]
        else:
            # Auto-register device
            device_id = qx(
                "INSERT OR IGNORE INTO devices (code, name, category, status) VALUES (?,?,?,?)",
                (device_code, f"ESP Device {device_code}", "Torque Turn System", "In Job")
            )
            if not device_id:
                dev2 = q1("SELECT id FROM devices WHERE code=?", (device_code,))
                device_id = dev2["id"] if dev2 else None

    ts = body.get("ts") or datetime.now().isoformat()
    torque      = body.get("torque")
    turns       = body.get("turns")
    rpm         = body.get("rpm")
    pressure    = body.get("pressure")
    temperature = body.get("temperature")
    weight      = body.get("weight")
    job_id      = body.get("job_id")

    # Save to DB
    rid = qx(
        "INSERT INTO scada_readings "        "(device_id, job_id, ts, torque, turns, rpm, pressure, temperature, weight, source) "        "VALUES (?,?,?,?,?,?,?,?,?,?)",
        (device_id, job_id, ts, torque, turns, rpm, pressure, temperature, weight, "esp32")
    )

    # Broadcast via WebSocket
    broadcast_data = {
        "type": "reading",
        "id": rid,
        "device_id": device_id,
        "device_code": device_code or "",
        "ts": ts,
        "torque": torque,
        "turns": turns,
        "rpm": rpm,
        "pressure": pressure,
        "temperature": temperature,
        "weight": weight,
        "joint_num": body.get("joint_num"),
        "job_id": job_id,
        "source": "esp32",
    }

    channel = str(device_id) if device_id else "all"
    import asyncio
    asyncio.create_task(ws_manager.broadcast(broadcast_data, channel))

    return JSONResponse({
        "ok": True,
        "id": rid,
        "ts": ts,
        "clients": ws_manager.client_count(),
    })


@app.get("/api/esp/status")
async def esp_status():
    """ESP32 can call this to verify connectivity."""
    return JSONResponse({
        "ok": True,
        "server": "TRS Platform",
        "version": APP_VERSION,
        "ts": datetime.now().isoformat(),
        "ws_clients": ws_manager.client_count(),
    })

# ══════════════════════════════════════════════════════════════
#  AI INSIGHTS
# ══════════════════════════════════════════════════════════════

@app.get("/ai", response_class=HTMLResponse)
async def ai_page(request: Request, job_id: int = None):
    user, redir = require_user(request)
    if redir: return redir
    if not rbac.can(user, "ai.view"):
        from fastapi.responses import HTMLResponse as HR
        return HR("""<meta http-equiv="refresh" content="2;url=/">
            <div style="font-family:sans-serif;background:#050d1a;color:#e2e8f0;padding:40px;text-align:center">
            🔒 Access denied — AI Insights requires Engineer role or above</div>""", 403)

    all_jobs = qa("""
        SELECT j.id, j.job_number, j.customer, m.total_joints, m.ft_mean, m.rerun_rate
        FROM jobs j JOIN mtt_job_summary m ON m.job_id=j.id
        ORDER BY j.created_at DESC
    """)

    analysis = ai_saved = None
    selected_job = None

    if job_id:
        selected_job = q1("""SELECT j.*, d.code device_code, d.name device_name
                              FROM jobs j LEFT JOIN devices d ON d.id=j.assigned_device_id
                              WHERE j.id=?""", (job_id,))
        ai_saved = q1("SELECT * FROM job_ai_insights WHERE job_id=?", (job_id,))

        if ai_saved and ai_saved.get("insights_json"):
            try: ai_saved["insights_json"] = json.loads(ai_saved["insights_json"])
            except Exception as e: logger.debug(f"Ignored non-critical error: {e}")

    # Fleet AI summary
    fleet_ai = qa("""
        SELECT jai.job_id, jai.performance_score, jai.stability_score,
               jai.anomaly_count, jai.torque_trend, jai.analyzed_at,
               j.job_number, j.customer
        FROM job_ai_insights jai JOIN jobs j ON j.id=jai.job_id
        ORDER BY jai.performance_score DESC
    """)

    return render(request, "ai.html", {
        "active": "ai", "all_jobs": all_jobs,
        "selected_job": selected_job, "selected_job_id": job_id,
        "ai_saved": ai_saved, "fleet_ai": fleet_ai,
    })


@app.post("/api/ai/analyze/{job_id}")
async def ai_analyze(job_id: int, request: Request):
    user = get_user(request)
    if not user:
        return JSONResponse({"ok": False, "error": "Not authorized"}, 401)

    summary = q1("""SELECT * FROM mtt_job_summary WHERE job_id=?""", (job_id,))
    joints  = qa("SELECT * FROM mtt_joint_data WHERE job_id=? ORDER BY joint_num", (job_id,))

    if not summary or not joints:
        return JSONResponse({"ok": False, "error": "No MTT data found for this job"}, 404)

    # Parse JSON fields
    stats_json = {}
    if summary.get("stats_json"):
        try: stats_json = json.loads(summary["stats_json"]) if isinstance(summary["stats_json"],str) else summary["stats_json"]
        except Exception as e: logger.debug(f"Ignored non-critical error: {e}")

    ft_vals = [j["final_torque"] for j in joints if j["final_torque"]]
    rpm_vals = [j["shoulder_rpm"] for j in joints if j["shoulder_rpm"]]
    turn_vals = [j["final_turns"] for j in joints if j["final_turns"]]
    dt_vals = [j["delta_torque"] for j in joints if j["delta_torque"]]

    def _mean(v): return round(sum(v)/len(v), 2) if v else 0
    def _std(v):
        if len(v) < 2: return 0
        m = _mean(v); return round(math.sqrt(sum((x-m)**2 for x in v)/(len(v)-1)), 2)
    def _cv(v): m = _mean(v); return round(_std(v)/m*100,1) if m else 0

    # Stability score (lower CV = more stable)
    stability = max(0, min(100, 100 - _cv(ft_vals) * 2))

    # Outlier detection (z-score > 2)
    mu, sig = _mean(ft_vals), _std(ft_vals)
    outliers = [j for j in joints if j["final_torque"] and sig > 0
                and abs(j["final_torque"] - mu) / sig > 2]

    # Rerun analysis
    reruns = [j for j in joints if j["is_rerun"]]
    rerun_rate = summary.get("rerun_rate", 0)

    # Performance score
    perf = 100.0
    perf -= min(25, len(outliers) * 3)
    perf -= min(20, rerun_rate * 2)
    perf = perf * 0.6 + stability * 0.4
    perf = max(0, min(100, perf))

    # Trend analysis (slope of torque)
    if len(ft_vals) >= 4:
        n = len(ft_vals)
        xs = list(range(n)); xm = (n-1)/2; ym = _mean(ft_vals)
        num = sum((xs[i]-xm)*(ft_vals[i]-ym) for i in range(n))
        den = sum((x-xm)**2 for x in xs)
        slope = num/den if den else 0
        trend = "increasing" if slope > 5 else "decreasing" if slope < -5 else "stable"
    else:
        trend = "stable"

    # Build insights
    insights = []
    max_t = stats_json.get("max_torque")
    min_t = stats_json.get("min_torque")
    opt_t = stats_json.get("opt_torque")

    if rerun_rate > 10:
        insights.append({"type":"critical","icon":"🔴","title":f"High Rerun Rate: {rerun_rate:.1f}%",
                         "detail":f"{summary['rerun_count']} reruns out of {summary['total_joints']} joints — check pipe condition and makeup specs"})
    elif rerun_rate > 5:
        insights.append({"type":"warning","icon":"🟡","title":f"Elevated Rerun Rate: {rerun_rate:.1f}%",
                         "detail":"Monitor rerun trends — could indicate tool wear or pipe issues"})
    else:
        insights.append({"type":"success","icon":"✅","title":f"Low Rerun Rate: {rerun_rate:.1f}%",
                         "detail":"Makeup operations are running cleanly"})

    if len(outliers) > 0:
        insights.append({"type":"warning","icon":"⚡","title":f"{len(outliers)} Outlier Joints",
                         "detail":f"Joints with torque deviation > 2σ: {', '.join(str(o['joint_num']) for o in outliers[:5])}"})

    if trend == "increasing":
        insights.append({"type":"warning","icon":"📈","title":"Torque Increasing Trend",
                         "detail":"Final torque is gradually rising — check tong calibration and pipe grade consistency"})
    elif trend == "decreasing":
        insights.append({"type":"info","icon":"📉","title":"Torque Decreasing Trend",
                         "detail":"Final torque trending down — verify tool condition and makeup parameters"})
    else:
        insights.append({"type":"success","icon":"📊","title":"Stable Torque Trend",
                         "detail":"Torque values are consistent across joints"})

    if stability > 80:
        insights.append({"type":"success","icon":"🎯","title":f"High Stability: {stability:.0f}%",
                         "detail":"Excellent consistency in makeup operations"})
    elif stability < 50:
        insights.append({"type":"critical","icon":"⚠️","title":f"Low Stability: {stability:.0f}%",
                         "detail":"High variability in torque readings — check equipment and operating conditions"})

    if rpm_vals:
        low_rpm = [j for j in joints if j.get("shoulder_rpm") and j["shoulder_rpm"] < 2.5]
        if low_rpm:
            insights.append({"type":"warning","icon":"🌀","title":f"{len(low_rpm)} Low RPM Joints",
                             "detail":"Low shoulder RPM may indicate tong slippage or speed control issues"})

    if opt_t and ft_vals:
        on_target = sum(1 for v in ft_vals if min_t and max_t and min_t <= v <= max_t)
        pct = round(on_target / len(ft_vals) * 100, 1)
        if pct >= 90:
            insights.append({"type":"success","icon":"🎯","title":f"{pct}% Joints Within Spec",
                             "detail":f"Most joints between {min_t}–{max_t} ft-lb (Optimum: {opt_t})"})
        else:
            insights.append({"type":"warning","icon":"📏","title":f"Only {pct}% Joints Within Spec",
                             "detail":f"Target: {min_t}–{max_t} ft-lb — {100-pct:.0f}% outside acceptable range"})

    # Save to DB
    qx("""
        INSERT INTO job_ai_insights
            (job_id, performance_score, stability_score, anomaly_count, torque_trend, insights_json, analyzed_at)
        VALUES (?,?,?,?,?,?,datetime('now'))
        ON CONFLICT(job_id) DO UPDATE SET
            performance_score=excluded.performance_score,
            stability_score=excluded.stability_score,
            anomaly_count=excluded.anomaly_count,
            torque_trend=excluded.torque_trend,
            insights_json=excluded.insights_json,
            analyzed_at=datetime('now')
    """, (job_id, round(perf,1), round(stability,1), len(outliers), trend,
          json.dumps(insights, ensure_ascii=False)))

    log_activity("ai_analyze", user["id"], "job", job_id, f"AI analysis: score={perf:.0f}")

    return JSONResponse({
        "ok": True,
        "performance_score": round(perf, 1),
        "stability_score": round(stability, 1),
        "anomaly_count": len(outliers),
        "torque_trend": trend,
        "insights": insights,
        "stats": {
            "ft_mean": _mean(ft_vals), "ft_std": _std(ft_vals),
            "rpm_mean": _mean(rpm_vals), "cv_pct": _cv(ft_vals),
            "rerun_rate": rerun_rate,
        }
    })


# ══════════════════════════════════════════════════════════════
#  USERS & SETTINGS
# ══════════════════════════════════════════════════════════════

@app.get("/users", response_class=HTMLResponse)
async def users_page(request: Request):
    user, redir = require_role(request, ["Admin", "Manager"])
    if redir: return redir

    users = qa("SELECT id, username, full_name, email, role, is_active, created_at FROM users ORDER BY created_at DESC")
    return render(request, "users.html", {
        "active": "users", "users": users, "roles": USER_ROLES,
    })


@app.post("/users/create")
async def user_create(request: Request,
                      username: str = Form(...),
                      full_name: str = Form(""),
                      email: str = Form(""),
                      role: str = Form("Field Operator"),
                      password: str = Form(...)):
    user, redir = require_role(request, ["Admin"])
    if redir: return redir

    existing = q1("SELECT id FROM users WHERE username=?", (username,))
    if existing:
        return RedirectResponse("/users?error=duplicate", status_code=303)

    pw = AuthService.hash_password(password)
    uid = qx("INSERT INTO users (username, full_name, email, role, password) VALUES (?,?,?,?,?)",
             (username, full_name, email, role, pw))
    log_activity("user_create", user["id"], "user", uid, f"Created user {username}")
    return RedirectResponse("/users", status_code=303)


@app.post("/users/{uid}/toggle")
async def user_toggle(uid: int, request: Request):
    user, redir = require_role(request, ["Admin"])
    if redir: return redir
    target = q1("SELECT * FROM users WHERE id=?", (uid,))
    if not target:
        return RedirectResponse("/users", status_code=303)
    qx("UPDATE users SET is_active=? WHERE id=?", (0 if target["is_active"] else 1, uid))
    return RedirectResponse("/users", status_code=303)


@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request):
    user, redir = require_role(request, ["Admin", "Manager"])
    if redir: return redir

    stats = {
        "jobs": q1("SELECT COUNT(*) c FROM jobs")["c"],
        "joints": q1("SELECT COALESCE(SUM(total_joints),0) c FROM mtt_job_summary")["c"],
        "devices": q1("SELECT COUNT(*) c FROM devices")["c"],
        "users": q1("SELECT COUNT(*) c FROM users")["c"],
        "files": q1("SELECT COUNT(*) c FROM job_data_files")["c"],
        "db_path": str(DB_PATH),
    }
    return render(request, "settings.html", {"active": "settings", "stats": stats})


# ══════════════════════════════════════════════════════════════
#  API STATUS & MISC
# ══════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════
#  QUICK START PAGE
# ══════════════════════════════════════════════════════════════

@app.get("/quickstart", response_class=HTMLResponse)
async def quickstart_page(request: Request):
    user, redir = require_user(request)
    if redir: return redir

    # System status checks
    job_count    = q1("SELECT COUNT(*) c FROM jobs")["c"]
    device_count = q1("SELECT COUNT(*) c FROM devices")["c"]
    user_count   = q1("SELECT COUNT(*) c FROM users WHERE is_active=1")["c"]
    mtt_count    = q1("SELECT COUNT(*) c FROM mtt_job_summary")["c"]
    scada_count  = q1("SELECT COUNT(*) c FROM scada_readings WHERE ts >= datetime('now','-1 hour')")["c"]

    checks = [
        {"id":"db",     "label":"Database",        "ok": True,            "detail":"SQLite connected"},
        {"id":"jobs",   "label":"Jobs",             "ok": job_count >= 0,  "detail": f"{job_count} jobs in system"},
        {"id":"devices","label":"Devices",          "ok": device_count > 0,"detail": f"{device_count} devices" if device_count > 0 else "No devices — add one first"},
        {"id":"esp",    "label":"ESP32 / SCADA",    "ok": scada_count > 0, "detail": f"{scada_count} readings last hour" if scada_count > 0 else "No live data — use Test Mode"},
        {"id":"mtt",    "label":"MTT Analysis",     "ok": mtt_count > 0,   "detail": f"{mtt_count} jobs analyzed" if mtt_count > 0 else "No MTT data — upload a PDF"},
        {"id":"fs",     "label":"Firestore",        "ok": fs_enabled(),    "detail":"Connected" if fs_enabled() else "Not configured (optional)"},
        {"id":"cl",     "label":"Cloudinary",       "ok": cl_enabled(),    "detail":"Connected" if cl_enabled() else "Not configured (optional)"},
    ]

    return render(request, "quickstart.html", {
        "active": "quickstart",
        "checks": checks,
        "job_count": job_count,
        "device_count": device_count,
        "mtt_count": mtt_count,
        "scada_count": scada_count,
        "esp_key": ESP_API_KEY,
    })


@app.get("/api/status")
async def api_status():
    """Health check endpoint — used by Render and ESP32."""
    try:
        jobs    = q1("SELECT COUNT(*) c FROM jobs")["c"]
        joints  = q1("SELECT COALESCE(SUM(total_joints),0) c FROM mtt_job_summary")["c"]
        devices = q1("SELECT COUNT(*) c FROM devices")["c"]
        db_ok   = True
    except Exception as e:
        jobs = joints = devices = 0
        db_ok = False
        logger.warning(f"Health check DB error: {e}")

    return JSONResponse({
        "app":     APP_NAME,
        "version": APP_VERSION,
        "status":  "ok",
        "db":      "ok" if db_ok else "error",
        "ts":      datetime.now().isoformat(),
        "jobs":    jobs,
        "joints":  joints,
        "devices": devices,
        "ws_clients": ws_manager.client_count(),
    })


@app.get("/api/jobs")
async def api_jobs(request: Request):
    user = get_user(request)
    if not user: return JSONResponse({"error": "Not authorized"}, 401)
    jobs = qa("""SELECT j.*, m.total_joints, m.ft_mean, m.rerun_rate
                 FROM jobs j LEFT JOIN mtt_job_summary m ON m.job_id=j.id
                 ORDER BY j.created_at DESC""")
    return JSONResponse(jobs)


@app.get("/api/jobs/{job_id}/joints")
async def api_joints(job_id: int, request: Request):
    user = get_user(request)
    if not user: return JSONResponse({"error": "Not authorized"}, 401)
    joints = qa("SELECT * FROM mtt_joint_data WHERE job_id=? ORDER BY joint_num ASC", (job_id,))
    return JSONResponse([dict(j) for j in joints])


@app.get("/api/devices")
async def api_devices(request: Request):
    user = get_user(request)
    if not user: return JSONResponse({"error": "Not authorized"}, 401)
    devices = qa("SELECT * FROM devices ORDER BY code")
    return JSONResponse(devices)


# ══════════════════════════════════════════════════════════════
#  404 handler
# ══════════════════════════════════════════════════════════════

from starlette.exceptions import HTTPException as StarletteHTTPException

@app.exception_handler(StarletteHTTPException)
async def http_exception_handler(request: Request, exc: StarletteHTTPException):
    if exc.status_code == 404:
        user = get_user(request)
        return HTMLResponse(f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<title>404 — TRS</title>
<style>body{{font-family:sans-serif;background:#0f172a;color:#fff;display:flex;align-items:center;
justify-content:center;height:100vh;margin:0;flex-direction:column;gap:1rem}}
h1{{font-size:5rem;margin:0;color:#f59e0b}} p{{color:#94a3b8}} a{{color:#3b82f6}}</style></head>
<body><h1>404</h1><p>Page not found</p><a href="/">← Back to Dashboard</a></body></html>""",
            status_code=404)
    raise exc


# ══════════════════════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
